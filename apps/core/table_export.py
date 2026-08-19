"""Download a list screen in whichever shape the person on the other end opens.

A screen points this at its ``ColumnSet`` and says how to fetch the rows; the
five formats come for free and all read the same builder, so a spreadsheet, a
document and a print sheet can never disagree about what was on screen.

    class GRNExportView(TableExportView):
        page = "inventory.grn"
        columns = GRN_COLUMNS
        filename = "goods-receipts"
        title = "Goods Receipts"

        def get_rows(self):
            return GRNListView(request=self.request, kwargs={}, args=()).export_rows()

Only the visible columns are written, so the file matches the table the person
was looking at rather than everything the screen could have shown.
"""

import csv
from datetime import date, datetime
from decimal import Decimal

from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.template.loader import render_to_string
from django.utils import timezone
from django.views.generic import View


class TableExportView(View):
    """Base for a screen's download menu. Subclasses set ``columns`` and rows."""

    columns = None
    filename = "export"
    title = "Export"
    doc_template = "components/table/export_doc.html"
    print_template = "components/table/export_print.html"

    FORMATS = ("xlsx", "csv", "pdf", "doc", "json")

    def get_rows(self):
        raise NotImplementedError

    def get(self, request, *args, **kwargs):
        columns = self.columns.exportable(request.session)
        rows_source = self.get_rows()

        # Built once, so every format below writes the same figures.
        header = [column.label for column in columns]
        rows = [[column.export(row) for column in columns] for row in rows_source]

        kind = request.GET.get("format", "csv")
        if kind not in self.FORMATS:
            kind = "csv"
        return getattr(self, f"_{kind}")(request, header, rows, columns)

    # ── formats ────────────────────────────────────────────────────────────
    def _csv(self, request, header, rows, columns):
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="{self.filename}.csv"'
        # Excel reads a CSV as the machine's own encoding unless the file says
        # otherwise, so the BOM is what keeps a supplier's name intact.
        response.write("﻿")
        writer = csv.writer(response)
        writer.writerow(header)
        writer.writerows(rows)
        return response

    def _xlsx(self, request, header, rows, columns):
        """A real workbook, not a CSV wearing an .xlsx name."""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        book = Workbook()
        sheet = book.active
        sheet.title = self.title[:31]
        sheet.append(header)

        heading = Font(bold=True, color="FFFFFF")
        band = PatternFill("solid", fgColor="1E293B")
        for cell in sheet[1]:
            cell.font = heading
            cell.fill = band
            cell.alignment = Alignment(vertical="center")

        for row in rows:
            sheet.append([self._native(value) for value in row])

        # Room to read, a frozen heading and a filter row: what anybody would do
        # to the sheet by hand the moment they opened it.
        for index, label in enumerate(header, start=1):
            widest = max([len(str(label))] + [len(str(row[index - 1])) for row in rows] or [0])
            sheet.column_dimensions[get_column_letter(index)].width = min(max(widest + 2, 10), 42)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{self.filename}.xlsx"'
        book.save(response)
        return response

    def _doc(self, request, header, rows, columns):
        """Word reads an HTML table as a document and keeps the formatting."""
        html = render_to_string(self.doc_template, self._paper(request, header, rows))
        response = HttpResponse(html, content_type="application/msword")
        response["Content-Disposition"] = f'attachment; filename="{self.filename}.doc"'
        return response

    def _pdf(self, request, header, rows, columns):
        """A print-ready page the browser saves as PDF.

        There is no PDF library here, and every browser already paginates a
        table properly -- page breaks, repeated headings and all.
        """
        return render(request, self.print_template, self._paper(request, header, rows))

    def _json(self, request, header, rows, columns):
        payload = [
            dict(zip([column.key for column in columns], [str(value) for value in row]))
            for row in rows
        ]
        response = JsonResponse({"columns": header, "rows": payload}, json_dumps_params={"indent": 2})
        response["Content-Disposition"] = f'attachment; filename="{self.filename}.json"'
        return response

    # ── shared bits ────────────────────────────────────────────────────────
    def _paper(self, request, header, rows):
        return {
            "title": self.title,
            "header": header,
            "rows": rows,
            "printed_by": request.user,
            "printed_on": timezone.localdate(),
        }

    @staticmethod
    def _native(value):
        """Keep a number a number and a date a date; everything else is text."""
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, (int, float, date, datetime)):
            return value
        return str(value)
