import csv
import io
import json

from datetime import date, datetime

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Count, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse_lazy
from django.utils import timezone
from django.views.generic import CreateView, DetailView, FormView, ListView, UpdateView, View

from decimal import Decimal, InvalidOperation
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit

from apps.core.constants import INV_PO_CANCEL_REASONS, INV_PO_CLOSE_SHORT_REASONS, INV_REVERSAL_REASONS, INVENTORY_KIND_PRODUCT, INVENTORY_KIND_SERVICE, INV_POS_STATUS_CHOICES, INV_PURCHASE_ORDER_STATUS_CHOICES, INV_TRANSACTION_TYPE_CHOICES, NO, RECORD_STATUS_CHOICES, STATUS_ACTIVE, STATUS_CREATED, STATUS_DRAFT, STATUS_FULLY_RECEIVED, STATUS_INACTIVE, STATUS_CANCELLED, STATUS_CLOSED_SHORT, STATUS_PARTIAL_RECEIVED, STATUS_POSTED, STATUS_RAISED, STATUS_REVERSED, YES
from apps.access_control.selectors import user_has_permission
from apps.core.table_export import TableExportView
from apps.core.mixins import PagePermissionRequiredMixin, PortalPermissionRequiredMixin, PrintContextMixin, SearchFilterPaginationMixin, SortableListMixin
from apps.finance.models import AccountVoucherLine, ChartOfAccount
from apps.finance.services import account_balances, account_ledger, create_customer_receivable_account, sync_supplier_opening_balance
from apps.finance.views import AuditSaveMixin

from .forms import PurchaseApprovalLimitForm, PurchaseBillForm, PurchaseOrderCancelForm, PurchaseOrderCloseShortForm, ReversalReasonForm, CustomerForm, InventoryClassForm, InventoryItemForm, InventoryItemImportForm, ManualTransactionForm, POSDetailForm, POSMasterForm, POSReturnDetailForm, POSReturnMasterForm, PurchaseOrderForm, PurchaseOrderItemForm, PurchaseReturnDetailForm, PurchaseReturnMasterForm, ReceivePOForm, UOMConversionForm, UOMForm, SupplierForm
from .models import PurchaseBill, PurchaseBillItem, Customer, CustomerLedger, InventoryClass, InventoryItem, ItemLedger, ManualTransaction, POSDetail, POSMaster, POSReturnDetail, POSReturnMaster, PurchaseMaster, PurchaseOrder, PurchaseOrderItem, PurchaseOrderItemReceived, PurchaseReturnDetail, PurchaseReturnMaster, Stock, UOM, UOMConversion, Supplier
from .purchase_board import COLUMNS, GRN_COLUMNS, TAB_ALL, TAB_UNBILLED, TABS, TAB_STATUSES, column_menu, decorate, export_columns, set_visible_columns, summarise, visible_columns
from .form_layout import EXTRA_FIELD_TYPES, add_extra_field, get_layout, read_extra_values, remove_extra_field, set_hidden
from .services import apportion_freight, approve_purchase_order, billable_receipts, can_reverse_bill, can_reverse_receipt, cancel_purchase_order, close_purchase_order_short, create_purchase_bill, needs_approval, purchase_order_approval_limit, reopen_purchase_order, reverse_purchase_bill, reverse_purchase_receipt, set_purchase_order_approval_limit, user_can_approve, amount_in_words, create_direct_purchase, create_direct_sale, create_purchase_order, finalize_manual_transaction, set_opening_stock, generate_transaction_id, next_direct_purchase_number, next_grn_number, next_purchase_order_number, next_sale_invoice_number, post_purchase_return, post_sale, post_sale_return, receive_purchase_order_item

User = get_user_model()


def uom_title(record):
    """The unit shown against a record, blank where it carries none.

    An item is allowed to have no unit, so every screen that prints one has to
    survive the empty case rather than reaching through a null relation.
    """
    return record.uom.title if record and record.uom_id else ""


def item_unit_options(item):
    """The units an item is actually handled in, and what each is worth in its own.

    Its own unit, the second one it is bought or issued in, and the other side
    of whatever conversion is set against it. An item with none configured
    returns nothing, and the caller falls back to the full list rather than
    leaving the operator with an empty dropdown.

    ``factor`` is how many base units one of that unit makes, so a screen can
    show what a quantity comes to in stock terms without asking the server.
    """
    units, seen = [], set()
    candidates = [item.uom, item.secondary_uom]
    if item.conversion_id:
        candidates += [item.conversion.uom_from, item.conversion.uom_to]
    for unit in candidates:
        if unit and unit.pk not in seen:
            seen.add(unit.pk)
            units.append({"id": unit.pk, "name": unit.title, "factor": unit_factor_to_base(item, unit)})
    return units


def unit_factor_to_base(item, unit):
    """How many of the item's own units one ``unit`` makes; 0 where none is set."""
    base = item.uom
    if not base or not unit:
        return 0
    if unit.pk == base.pk:
        return 1
    down = UOMConversion.objects.filter(uom_from=base, uom_to=unit, status=STATUS_ACTIVE).first()
    if down and down.conversion_factor:
        return float(1 / down.conversion_factor)
    up = UOMConversion.objects.filter(uom_from=unit, uom_to=base, status=STATUS_ACTIVE).first()
    if up and up.conversion_factor:
        return float(up.conversion_factor)
    return 0


class InventoryListMixin(SearchFilterPaginationMixin, PagePermissionRequiredMixin):
    pass


class InventoryManageMixin(AuditSaveMixin, PagePermissionRequiredMixin):
    pass


class BaseSimpleListView(InventoryListMixin, ListView):
    template_name = "inventory/simple_list.html"
    context_object_name = "records"
    extra_context = {}

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(self.extra_context)
        return context


class InventoryClassListView(BaseSimpleListView):
    """Categories on the left, the items filed under one of them on the right.

    Same shape as the units screen: nothing here navigates, every exchange
    swaps the two panes. "Items not in any category" is a row in the list but
    not a record -- it stands for item_class being null.
    """

    page = "inventory.classes"
    model = InventoryClass
    template_name = "inventory/class_list.html"
    queryset = InventoryClass.objects.order_by("title")
    # A picker in a scrolling panel, not a page of results.
    paginate_by = 100
    search_fields = ("title", "class_code")
    extra_context = {"title": "Item Categories", "active_tab": "category"}

    # The left row that stands for "no category at all".
    UNFILED = "none"

    def _is_ajax(self):
        return self.request.headers.get("X-Requested-With") == "XMLHttpRequest"

    def get_queryset(self):
        # The count beside each category is what the left column shows.
        return super().get_queryset().annotate(item_count=Count("inventoryitem", distinct=True))

    def get_selected(self):
        """The category on show, or the sentinel for the unfiled row.

        Returns (selected_class, is_unfiled). Nothing chosen means the unfiled
        row, so the screen always opens on something.
        """
        raw = self.request.GET.get("selected_class") or self.request.POST.get("selected_class") or ""
        if raw in ("", self.UNFILED):
            return None, True
        return InventoryClass.objects.filter(pk=raw).first(), False

    def get_items(self, selected_class, is_unfiled, search=""):
        items = InventoryItem.objects.select_related("stock").order_by("item_name")
        items = items.filter(item_class__isnull=True) if is_unfiled else items.filter(item_class=selected_class)
        if search:
            items = items.filter(Q(item_name__icontains=search) | Q(code__icontains=search))
        return items

    @staticmethod
    def _decorate_items(rows):
        """Stock value per row, so the table does no arithmetic of its own."""
        rows = list(rows)
        for row in rows:
            stock = getattr(row, "stock", None)
            quantity = stock.current_quantity if stock else Decimal("0.0000")
            price = stock.current_price if stock else Decimal("0.00")
            row.stock_quantity = quantity
            row.stock_value = (quantity * price).quantize(Decimal("0.01"))
        return rows

    def _pane_context(self, selected_class, is_unfiled):
        self.object_list = self.get_queryset()
        context = self.get_context_data()
        context["selected_class"] = selected_class
        context["is_unfiled"] = is_unfiled
        context["item_search"] = self.request.GET.get("item_q", "").strip()
        context["items"] = self._decorate_items(
            self.get_items(selected_class, is_unfiled, context["item_search"])
        )
        return context

    def _fragments(self, selected_class, is_unfiled):
        context = self._pane_context(selected_class, is_unfiled)
        return {
            "rows_html": render_to_string("inventory/_class_rows.html", context, request=self.request),
            "detail_html": render_to_string("inventory/_class_detail.html", context, request=self.request),
            "selected_class": "" if is_unfiled else selected_class.pk,
        }

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["unfiled_count"] = InventoryItem.objects.filter(item_class__isnull=True).count()
        context["category_form"] = InventoryClassForm()
        context.setdefault("selected_class", None)
        context.setdefault("is_unfiled", True)
        return context

    # ---- AJAX -----------------------------------------------------------
    def get(self, request, *args, **kwargs):
        if not self._is_ajax():
            return super().get(request, *args, **kwargs)
        selected_class, is_unfiled = self.get_selected()
        if request.GET.get("mode") == "picker":
            return JsonResponse({"picker_html": self._picker_html(selected_class, is_unfiled)})
        return JsonResponse(self._fragments(selected_class, is_unfiled))

    def _picker_html(self, selected_class, is_unfiled):
        """Rows for the Select Items dialog: everything not already filed here."""
        search = self.request.GET.get("pick_q", "").strip()
        items = InventoryItem.objects.select_related("item_class", "stock").order_by("item_name")
        items = items.filter(item_class__isnull=False) if is_unfiled else items.exclude(item_class=selected_class)
        if search:
            items = items.filter(Q(item_name__icontains=search) | Q(code__icontains=search))
        context = {"items": self._decorate_items(items[:200])}
        return render_to_string("inventory/_class_picker.html", context, request=self.request)

    def _refused(self, message, status):
        if self._is_ajax():
            return JsonResponse({"ok": False, "errors": {"__all__": [message]}}, status=status)
        messages.error(self.request, message)
        return redirect("inventory:class_list")

    def _saved(self, selected_class, is_unfiled, message):
        if self._is_ajax():
            payload = self._fragments(selected_class, is_unfiled)
            payload.update(ok=True, message=message)
            return JsonResponse(payload)
        messages.success(self.request, message)
        base = str(reverse_lazy("inventory:class_list"))
        return redirect(base if is_unfiled else f"{base}?selected_class={selected_class.pk}")

    def _rejected(self, form):
        if self._is_ajax():
            errors = {field: [str(e) for e in errs] for field, errs in form.errors.items()}
            return JsonResponse({"ok": False, "errors": errors}, status=400)
        for errs in form.errors.values():
            for error in errs:
                messages.error(self.request, error)
        return redirect("inventory:class_list")

    def post(self, request, *args, **kwargs):
        if not user_has_permission(request.user, f"{self.page}.add"):
            return self._refused("You cannot change categories.", 403)

        action = request.POST.get("action")
        if action == "category":
            return self._save_category(request)
        if action == "category_delete":
            return self._delete_category(request)
        if action == "move":
            return self._move_items(request)
        return self._refused("Unknown action.", 400)

    def _save_category(self, request):
        category_id = request.POST.get("category_id")
        instance = InventoryClass.objects.filter(pk=category_id).first() if category_id else None
        if category_id and not instance:
            return self._refused("That category no longer exists.", 404)

        data = request.POST.copy()
        # The dialog asks for a name only; the code is derived the same way the
        # item form derives it, so two screens never invent different codes.
        if not data.get("class_code"):
            data["class_code"] = instance.class_code if instance else InventoryItemForm._next_class_code(data.get("title", ""))
        data.setdefault("status", STATUS_ACTIVE)

        form = InventoryClassForm(data, instance=instance)
        if form.is_valid():
            category = form.save(commit=False)
            category.created_by = category.created_by or request.user
            category.updated_by = request.user
            category.save()
            return self._saved(category, False, "Category updated." if instance else "Category created.")
        return self._rejected(form)

    def _delete_category(self, request):
        if not user_has_permission(request.user, f"{self.page}.delete"):
            return self._refused("You cannot delete categories.", 403)

        category = InventoryClass.objects.filter(pk=request.POST.get("category_id")).first()
        if not category:
            return self._refused("That category no longer exists.", 404)

        # An item is moved out of a category, never deleted with it, so a
        # category still holding items is refused until it is emptied.
        held = InventoryItem.objects.filter(item_class=category).count()
        if held:
            return self._refused(
                f"{category.title} still holds {held} item{'s' if held > 1 else ''}. Move them out first.", 400
            )

        category.soft_delete(user=request.user)
        return self._saved(None, True, f"{category.title} deleted.")

    def _move_items(self, request):
        """File the ticked items under the open category.

        item_class holds one category, so a move is a reassignment. Items that
        already sit somewhere else are only touched when the operator says so,
        which is what the tick box on the dialog asks.
        """
        selected_class, is_unfiled = self.get_selected()
        if not is_unfiled and not selected_class:
            return self._refused("That category no longer exists.", 404)

        ids = request.POST.getlist("item_ids")
        if not ids:
            return self._refused("Pick at least one item.", 400)

        items = InventoryItem.objects.filter(pk__in=ids)
        take_filed = request.POST.get("remove_existing") == "1"
        if not take_filed:
            already = items.filter(item_class__isnull=False).count()
            items = items.filter(item_class__isnull=True)
            if not items.exists():
                return self._refused(
                    f"{already} of those already sit in another category. "
                    "Tick 'Remove selected items from existing category' to move them.", 400
                )

        moved = 0
        for item in items:
            item.item_class = None if is_unfiled else selected_class
            item.updated_by = request.user
            item.save(update_fields=["item_class", "updated_by", "updated_at"])
            moved += 1

        where = "no category" if is_unfiled else selected_class.title
        return self._saved(selected_class, is_unfiled, f"{moved} item{'s' if moved > 1 else ''} moved to {where}.")


class InventoryClassToggleStatusView(InventoryManageMixin, View):
    page = "inventory.classes"
    action = "edit"
    def post(self, request, pk):
        record = get_object_or_404(InventoryClass, pk=pk)
        record.status = STATUS_INACTIVE if record.status == STATUS_ACTIVE else STATUS_ACTIVE
        record.updated_by = request.user
        record.save(update_fields=["status", "updated_by", "updated_at"])
        return redirect(request.META.get("HTTP_REFERER") or reverse_lazy("inventory:class_list"))


class UOMToggleStatusView(InventoryManageMixin, View):
    page = "inventory.uoms"
    action = "edit"
    def post(self, request, pk):
        record = get_object_or_404(UOM, pk=pk)
        record.status = STATUS_INACTIVE if record.status == STATUS_ACTIVE else STATUS_ACTIVE
        record.updated_by = request.user
        record.save(update_fields=["status", "updated_by", "updated_at"])
        return redirect(request.META.get("HTTP_REFERER") or reverse_lazy("inventory:uom_list"))


class InventoryClassCreateView(InventoryManageMixin, CreateView):
    page = "inventory.classes"
    model = InventoryClass
    form_class = InventoryClassForm
    template_name = "inventory/simple_form.html"
    success_url = reverse_lazy("inventory:class_list")
    success_message = "Inventory class saved."
    extra_context = {"title": "Item Category"}


class InventoryClassUpdateView(InventoryClassCreateView, UpdateView):
    success_message = "Inventory class updated."


class UOMListView(BaseSimpleListView):
    page = "inventory.uoms"
    model = UOM
    template_name = "inventory/uom_list.html"
    queryset = UOM.objects.order_by("title")
    # The list is a picker in a scrolling panel, not a page of results, so it
    # holds a full alphabet rather than ten rows.
    paginate_by = 100
    search_fields = ("title", "code")
    filter_fields = {"status": "status"}
    extra_context = {"title": "Units of Measure", "create_url": reverse_lazy("inventory:uom_create"), "edit_url_name": "inventory:uom_update", "active_tab": "units", "columns": [("Title", "title"), ("Code", "code"), ("Status", "get_status_display")]}

    def get_filter_specs(self):
        return [{"name": "status", "label": "All statuses", "choices": RECORD_STATUS_CHOICES, "value": self.request.GET.get("status", "")}] 

    def get_selected_uom(self):
        selected_uom_id = self.request.GET.get("selected_uom") or self.request.POST.get("selected_uom")
        if not selected_uom_id:
            return None
        return UOM.objects.filter(pk=selected_uom_id).first()

    def get_conversions(self, selected_uom):
        """Every conversion held against the selected unit.

        A unit can be measured more than one way -- a bag is 20kg on one line
        and 50kg on another -- so the panel lists them rather than holding one.
        """
        if not selected_uom:
            return UOMConversion.objects.none()
        return UOMConversion.objects.filter(uom_from=selected_uom).select_related("uom_from", "uom_to")

    def get_conversion_form(self, selected_uom, instance=None, data=None):
        form = UOMConversionForm(data=data, instance=instance)
        # The base is the unit already open on the left; only the other side is
        # a choice, and it can never be the same unit.
        form.fields["uom_from"].queryset = UOM.objects.filter(pk=selected_uom.pk) if selected_uom else UOM.objects.none()
        form.fields["uom_to"].queryset = UOM.objects.exclude(pk=selected_uom.pk) if selected_uom else UOM.objects.none()
        if selected_uom and not form.is_bound:
            form.initial.setdefault("uom_from", selected_uom.pk)
        return form

    # ---- AJAX plumbing ---------------------------------------------------
    # The screen never navigates: it asks for the two panes and swaps them in.
    # Every handler answers JSON to an AJAX caller; a plain request still gets
    # the whole page, which is what the first load and a reload go through.

    def _is_ajax(self):
        return self.request.headers.get("X-Requested-With") == "XMLHttpRequest"

    def _fragments(self, selected_uom, **extra):
        self.object_list = self.get_queryset()
        context = self.get_context_data(**extra)
        context["selected_uom"] = selected_uom
        context["uom_conversions"] = self.get_conversions(selected_uom)
        payload = {
            "rows_html": render_to_string("inventory/_uom_rows.html", context, request=self.request),
            "detail_html": render_to_string("inventory/_uom_detail.html", context, request=self.request),
            "selected_uom": selected_uom.pk if selected_uom else "",
        }
        return payload

    @staticmethod
    def _errors(form):
        return {field: [str(e) for e in errors] for field, errors in form.errors.items()}

    def get(self, request, *args, **kwargs):
        if not self._is_ajax():
            return super().get(request, *args, **kwargs)
        return JsonResponse(self._fragments(self.get_selected_uom()))

    def _redirect_to_unit(self, selected_uom):
        base = str(reverse_lazy("inventory:uom_list"))
        return redirect(f"{base}?selected_uom={selected_uom.pk}" if selected_uom else base)

    def _saved(self, selected_uom, message):
        if self._is_ajax():
            payload = self._fragments(selected_uom)
            payload.update(ok=True, message=message)
            return JsonResponse(payload)
        messages.success(self.request, message)
        return self._redirect_to_unit(selected_uom)

    def _not_found(self, message):
        return self._refused(message, status=404)

    def _forbidden(self, message):
        return self._refused(message, status=403)

    def _rejected_message(self, message):
        return self._refused(message, status=400)

    def _refused(self, message, status):
        if self._is_ajax():
            return JsonResponse({"ok": False, "errors": {"__all__": [message]}}, status=status)
        messages.error(self.request, message)
        return redirect("inventory:uom_list")

    def _rejected(self, form):
        # Only the modal that was submitted needs to know, and it is on screen
        # already, so the errors go back on their own.
        if self._is_ajax():
            return JsonResponse({"ok": False, "errors": self._errors(form)}, status=400)
        for errors in form.errors.values():
            for error in errors:
                messages.error(self.request, error)
        return self._redirect_to_unit(self.get_selected_uom())

    def post(self, request, *args, **kwargs):
        # The list itself only needs index rights; writing needs add.
        if not user_has_permission(request.user, f"{self.page}.add"):
            return self._forbidden("You cannot change units.")

        action = request.POST.get("action")
        if action == "unit":
            return self._save_unit(request)
        if action == "unit_delete":
            return self._delete_unit(request)
        if action == "conversion_delete":
            return self._delete_conversion(request)
        return self._save_conversion(request)

    def _save_unit(self, request):
        # A posted id edits that unit; without one this is a new one.
        unit_id = request.POST.get("unit_id")
        instance = UOM.objects.filter(pk=unit_id).first() if unit_id else None
        if unit_id and not instance:
            return self._not_found("That unit no longer exists.")

        form = UOMForm(request.POST, instance=instance)
        if form.is_valid():
            unit = form.save(commit=False)
            unit.created_by = unit.created_by or request.user
            unit.updated_by = request.user
            unit.save()
            # A unit saved from here becomes the one on show.
            return self._saved(unit, "Unit updated." if instance else "Unit saved.")
        return self._rejected(form)

    @staticmethod
    def _usage(record, ignore=()):
        """Everything in the database still pointing at this record.

        Walked off the model's own relations rather than a hand-written list,
        so a table added later is counted without anyone remembering to come
        back here. A figure that was measured in a unit has to keep reading
        back the same way, so anything still referenced is never deleted.
        """
        found = []
        for relation in record._meta.related_objects:
            model = relation.related_model
            if model in ignore:
                continue
            manager = getattr(model, "objects", model._default_manager)
            count = manager.filter(**{relation.field.name: record}).count()
            if count:
                label = model._meta.verbose_name if count == 1 else model._meta.verbose_name_plural
                found.append(f"{count} {label}")
        return found

    def _delete_unit(self, request):
        """Retire a unit, provided nothing anywhere is measured in it."""
        if not user_has_permission(request.user, f"{self.page}.delete"):
            return self._forbidden("You cannot delete units.")

        unit = UOM.objects.filter(pk=request.POST.get("unit_id")).first()
        if not unit:
            return self._not_found("That unit no longer exists.")

        # Its own conversions are not a reason to stop -- they go with it --
        # but a conversion that something else uses is, so they are checked in
        # their own right below.
        blockers = self._usage(unit, ignore=(UOMConversion,))
        own_conversions = UOMConversion.objects.filter(Q(uom_from=unit) | Q(uom_to=unit))
        for conversion in own_conversions:
            blockers += self._usage(conversion)
        if blockers:
            return self._rejected_message(
                f"{unit.title} cannot be deleted: it is still used by {', '.join(blockers)}. "
                "Deactivate it instead, so past figures keep reading back."
            )

        for conversion in own_conversions:
            conversion.soft_delete(user=request.user)
        unit.soft_delete(user=request.user)

        selected_uom = self.get_selected_uom()
        if selected_uom and selected_uom.pk == unit.pk:
            selected_uom = None  # the pane was showing what just went away
        return self._saved(selected_uom, f"{unit.title} deleted.")

    def _delete_conversion(self, request):
        if not user_has_permission(request.user, f"{self.page}.delete"):
            return self._forbidden("You cannot delete conversions.")

        selected_uom = self.get_selected_uom()
        conversion = self.get_conversions(selected_uom).filter(pk=request.POST.get("conversion_id")).first()
        if not conversion:
            return self._not_found("That conversion no longer exists.")

        blockers = self._usage(conversion)
        if blockers:
            return self._rejected_message(
                f"This conversion cannot be deleted: it is still used by {', '.join(blockers)}. "
                "Deactivate it instead, so past figures keep reading back."
            )

        conversion.soft_delete(user=request.user)
        return self._saved(selected_uom, "Conversion deleted.")

    def _save_conversion(self, request):
        selected_uom = self.get_selected_uom()
        if not selected_uom:
            return self._rejected_message("Select a unit first.")

        # A posted id edits that row; without one this is a new conversion, so
        # a unit can carry as many as it needs.
        conversion_id = request.POST.get("conversion_id")
        instance = self.get_conversions(selected_uom).filter(pk=conversion_id).first() if conversion_id else None

        form = self.get_conversion_form(selected_uom, instance=instance, data=request.POST)
        if form.is_valid():
            conversion = form.save(commit=False)
            conversion.created_by = conversion.created_by or request.user
            conversion.updated_by = request.user
            conversion.save()
            return self._saved(selected_uom, "Conversion updated." if instance else "Conversion saved.")
        return self._rejected(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        selected_uom = self.get_selected_uom()
        context["selected_uom"] = selected_uom
        context["uom_conversions"] = self.get_conversions(selected_uom)
        context["unit_form"] = UOMForm()
        # The secondary-unit picker is built once and reused for every unit, so
        # it carries the whole list and hides the base unit in script.
        context["all_units"] = UOM.objects.order_by("title")
        return context


class UOMCreateView(InventoryManageMixin, CreateView):
    page = "inventory.uoms"
    model = UOM
    form_class = UOMForm
    template_name = "inventory/simple_form.html"
    success_url = reverse_lazy("inventory:uom_list")
    success_message = "UOM saved."
    extra_context = {"title": "UOM"}


class UOMUpdateView(UOMCreateView, UpdateView):
    success_message = "UOM updated."


class UOMConversionListView(BaseSimpleListView):
    page = "inventory.uom_conversions"
    model = UOMConversion
    queryset = UOMConversion.objects.select_related("uom_from", "uom_to").order_by("uom_from__title")
    search_fields = ("uom_from__title", "uom_to__title")
    filter_fields = {"status": "status"}
    extra_context = {"title": "UOM Conversions", "create_url": reverse_lazy("inventory:conversion_create"), "edit_url_name": "inventory:conversion_update", "status_toggle_url_name": "inventory:conversion_toggle_status", "columns": [("From", "uom_from"), ("To", "uom_to"), ("Factor", "conversion_factor"), ("Status", "status_toggle")]}

    def get_filter_specs(self):
        return [{"name": "status", "label": "All statuses", "choices": RECORD_STATUS_CHOICES, "value": self.request.GET.get("status", "")}] 


class UOMConversionToggleStatusView(InventoryManageMixin, View):
    page = "inventory.uom_conversions"
    action = "edit"
    def post(self, request, pk):
        record = get_object_or_404(UOMConversion, pk=pk)
        record.status = STATUS_INACTIVE if record.status == STATUS_ACTIVE else STATUS_ACTIVE
        record.updated_by = request.user
        record.save(update_fields=["status", "updated_by", "updated_at"])
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            # The units screen refreshes its own panes, so it only needs the word.
            return JsonResponse({"ok": True, "status": record.status, "message": "Conversion activated." if record.status == STATUS_ACTIVE else "Conversion deactivated."})
        return redirect(request.META.get("HTTP_REFERER") or reverse_lazy("inventory:conversion_list"))


class UOMConversionCreateView(InventoryManageMixin, CreateView):
    page = "inventory.uom_conversions"
    model = UOMConversion
    form_class = UOMConversionForm
    template_name = "inventory/simple_form.html"
    success_url = reverse_lazy("inventory:conversion_list")
    success_message = "UOM conversion saved."
    extra_context = {"title": "UOM Conversion"}


class UOMConversionUpdateView(UOMConversionCreateView, UpdateView):
    success_message = "UOM conversion updated."


class SupplierListView(SortableListMixin, BaseSimpleListView):
    """Suppliers, each row opening onto what the business has bought from them."""

    page = "inventory.suppliers"
    model = Supplier
    template_name = "inventory/supplier_list.html"
    # Newest first: a supplier just added is the one being looked for.
    queryset = Supplier.objects.select_related("city").order_by("-id")
    search_fields = ("name", "code", "email", "tel1")
    sort_fields = {"name": "name", "code": "code", "city": "city__title", "status": ("status", "name"), "added": "-id"}
    # Rolled up per row after the query, so the database cannot order by it.
    python_sort_fields = {"payable": "payable_balance"}
    default_sort = "added"
    filter_fields = {"status": "status"}
    extra_context = {"title": "Suppliers", "create_url": reverse_lazy("inventory:supplier_create"), "edit_url_name": "inventory:supplier_update", "status_toggle_url_name": "inventory:supplier_toggle_status"}

    def get_filter_specs(self):
        return [{"name": "status", "label": "All statuses", "choices": RECORD_STATUS_CHOICES, "value": self.request.GET.get("status", "")}]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        suppliers = list(context.get("records") or [])
        if not suppliers:
            return context

        zero = Decimal("0.00")
        # One query per kind for the whole page rather than per expanded row:
        # the panels are open by default, so every row's figures are needed.
        purchases = {
            row["supplier_id"]: row
            for row in PurchaseMaster.objects.filter(supplier__in=suppliers)
            .values("supplier_id")
            .annotate(total=Sum("total_amount"), count=Count("id"))
        }
        returns = {
            row["supplier_id"]: row
            for row in PurchaseReturnMaster.objects.filter(supplier__in=suppliers)
            .values("supplier_id")
            .annotate(total=Sum("returned_amount"), count=Count("id"))
        }
        orders = {
            row["supplier_id"]: row["count"]
            for row in PurchaseOrder.objects.filter(supplier__in=suppliers).values("supplier_id").annotate(count=Count("id"))
        }
        # A supplier's payable account is named after them, so the ledger the
        # panel links to is found the same way the posting created it.
        payable_codes = dict(
            ChartOfAccount.objects.filter(title__in=[supplier.name for supplier in suppliers])
            .values_list("title", "code")
        )
        balances = account_balances()

        # Every supplier's ledger in one pass: the panels are ledgers, and one
        # query per open panel would be a query per row.
        entries = {}
        if payable_codes:
            lines = (
                AccountVoucherLine.objects.filter(account_no__in=payable_codes.values())
                .select_related("voucher")
                .order_by("voucher_date", "voucher_no", "line_number")
            )
            for line in lines:
                entries.setdefault(line.account_no, []).append(line)

        def ledger_rows(code):
            """Payables are credit-natured: a purchase raises the balance, a payment lowers it."""
            running = (balances.get(code) or {}).get("opening") or zero
            rows = []
            for line in entries.get(code, []):
                debit = line.debit_amount or zero
                credit = line.credit_amount or zero
                running += credit - debit
                rows.append({"line": line, "debit": debit, "credit": credit, "balance": running})
            return rows[::-1][:8]  # newest first, the last few dealings

        for supplier in suppliers:
            bought = purchases.get(supplier.id) or {}
            sent_back = returns.get(supplier.id) or {}
            supplier.purchase_total = bought.get("total") or zero
            supplier.purchase_count = bought.get("count") or 0
            supplier.return_total = sent_back.get("total") or zero
            supplier.return_count = sent_back.get("count") or 0
            supplier.order_count = orders.get(supplier.id, 0)
            supplier.payable_code = payable_codes.get(supplier.name, "")
            supplier.payable_balance = (balances.get(supplier.payable_code) or {}).get("closing") or zero
            supplier.ledger_rows = ledger_rows(supplier.payable_code) if supplier.payable_code else []

        # The payable column is computed above, so its sort happens here.
        context["records"] = self.sort_rows(suppliers)
        # The tiles report the whole filtered set, not just the page in front of
        # you: "10 suppliers" on page 1 of 4 would be a lie.
        all_suppliers = self.get_queryset()
        context["supplier_count"] = all_suppliers.count()
        context["purchased_total"] = (
            PurchaseMaster.objects.filter(supplier__in=all_suppliers).aggregate(total=Sum("total_amount"))["total"] or zero
        )
        all_codes = ChartOfAccount.objects.filter(
            title__in=all_suppliers.values_list("name", flat=True), is_group=False
        ).values_list("code", flat=True)
        context["payable_total"] = sum(((balances.get(code) or {}).get("closing") or zero for code in all_codes), zero)
        return context


class SupplierDetailView(PagePermissionRequiredMixin, DetailView):
    """Everything on file for one supplier, with their ledger underneath."""

    page = "inventory.suppliers"
    model = Supplier
    template_name = "inventory/supplier_detail.html"
    context_object_name = "supplier"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        supplier = self.object
        zero = Decimal("0.00")

        bought = PurchaseMaster.objects.filter(supplier=supplier).aggregate(total=Sum("total_amount"), count=Count("id"))
        sent_back = PurchaseReturnMaster.objects.filter(supplier=supplier).aggregate(total=Sum("returned_amount"), count=Count("id"))
        context["purchase_total"] = bought["total"] or zero
        context["purchase_count"] = bought["count"] or 0
        context["return_total"] = sent_back["total"] or zero
        context["return_count"] = sent_back["count"] or 0
        context["order_count"] = PurchaseOrder.objects.filter(supplier=supplier).count()
        context["recent_receipts"] = PurchaseMaster.objects.filter(supplier=supplier).order_by("-id")[:10]

        # The supplier's payable account is named after them, the same way the
        # posting created it, so the ledger here is the ledger the books hold.
        account = ChartOfAccount.objects.filter(title=supplier.name, is_group=False).first()
        context["payable_code"] = account.code if account else ""
        ledger = account_ledger(account.code) if account else None
        context["ledger"] = ledger
        context["payable_balance"] = ledger["closing"] if ledger else zero
        return context


class SupplierToggleStatusView(InventoryManageMixin, View):
    page = "inventory.suppliers"
    action = "edit"
    def post(self, request, pk):
        record = get_object_or_404(Supplier, pk=pk)
        record.status = STATUS_INACTIVE if record.status == STATUS_ACTIVE else STATUS_ACTIVE
        record.updated_by = request.user
        record.save(update_fields=["status", "updated_by", "updated_at"])
        return redirect(request.META.get("HTTP_REFERER") or reverse_lazy("inventory:supplier_list"))


class EmbeddedCreateMixin:
    """A create screen that can also be opened in a modal on another screen.

    The whole form is framed rather than a thinner copy of it being written for
    the modal, so a record added mid-entry is the same record, with the same
    validation, as one added from its own menu. On save the frame tells the
    screen underneath what was created and that screen closes the modal.
    """

    # What the saved record is called in the message the frame posts up.
    embed_message_type = ""

    def embed_payload(self, obj):
        raise NotImplementedError

    def is_embedded(self):
        return self.request.GET.get("embed") == "1"

    def dispatch(self, request, *args, **kwargs):
        response = super().dispatch(request, *args, **kwargs)
        if self.is_embedded():
            # The site denies framing outright; these screens are allowed to be
            # framed by the portal itself, and by nothing else.
            response.xframe_options_exempt = True
            response["X-Frame-Options"] = "SAMEORIGIN"
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.is_embedded():
            context["embed"] = True
            context["embed_layout"] = "layouts/embed.html"
        return context

    def embed_saved_response(self):
        # A redirect here would only reload the form inside the frame.
        return render(self.request, "inventory/_embed_saved.html", {
            "message_type": self.embed_message_type,
            "payload_json": json.dumps(self.embed_payload(self.object)),
        })


class SupplierCreateView(EmbeddedCreateMixin, InventoryManageMixin, CreateView):
    page = "inventory.suppliers"
    model = Supplier
    form_class = SupplierForm
    template_name = "inventory/supplier_form.html"
    success_url = reverse_lazy("inventory:supplier_list")
    success_message = "Supplier saved."
    embed_message_type = "supplier:saved"
    extra_context = {
        "title": "Supplier",
        # Name, phone, address and the credit block are placed by hand; the rest
        # are grouped behind their own tabs so the first screen stays short.
        "registration_fields": ("code", "ntn_number", "sale_tax_num", "web_url"),
        "extra_fields": ("fax", "tel2", "status", "supplier_current_status", "remarks"),
    }

    def embed_payload(self, obj):
        return {"id": obj.pk, "name": obj.name, "balance": str(obj.opening_balance or "")}

    def form_valid(self, form):
        response = super().form_valid(form)
        # The opening balance belongs in the ledger, not only on the master.
        sync_supplier_opening_balance(supplier=self.object, user=self.request.user)
        if self.is_embedded():
            return self.embed_saved_response()
        return response

    def get_success_url(self):
        # "Save & New" is for entering suppliers in a run: it comes straight back
        # to an empty form instead of the list.
        if "save_and_new" in self.request.POST:
            return reverse_lazy("inventory:supplier_create")
        return super().get_success_url()


class SupplierUpdateView(SupplierCreateView, UpdateView):
    success_message = "Supplier updated."


class ItemStockListMixin(InventoryListMixin):
    """Items with the stock figures that used to live on their own page.

    Stock is one row per item, so the two are read together rather than kept as
    separate screens; the item list is the only place either is now shown.
    """

    page = "inventory.items"
    model = InventoryItem
    context_object_name = "records"
    queryset = InventoryItem.objects.select_related("uom", "item_class", "stock").order_by("item_name")
    # The stock row carries its own copy of the code and name, and is what a
    # search for a stocked item is likely typed against.
    search_fields = ("item_name", "code", "item_bar_code", "stock__item_code", "stock__item_name")
    filter_fields = {"status": "status", "item_class": "item_class_id"}

    def item_kind(self):
        """Which tab is being viewed. Products unless Services is asked for."""
        return INVENTORY_KIND_SERVICE if self.request.GET.get("kind") == INVENTORY_KIND_SERVICE else INVENTORY_KIND_PRODUCT

    def get_queryset(self):
        return super().get_queryset().filter(item_kind=self.item_kind())

    def get_filter_specs(self):
        class_choices = list(InventoryClass.objects.order_by("title").values_list("id", "title"))
        return [
            {"name": "status", "label": "All statuses", "choices": RECORD_STATUS_CHOICES, "value": self.request.GET.get("status", "")},
            {"name": "item_class", "label": "All classes", "choices": class_choices, "value": self.request.GET.get("item_class", "")},
        ]

    @staticmethod
    def _decorate(rows):
        """Attach the per-row stock value and return the (quantity, value) totals."""
        total_quantity = Decimal("0.0000")
        total_value = Decimal("0.00")
        for row in rows:
            stock = getattr(row, "stock", None)
            quantity = stock.current_quantity if stock else Decimal("0.0000")
            # Valued at the current price, the same figure the Inventory control
            # account is reconciled against.
            price = stock.current_price if stock else row.price
            row.stock_value = (quantity * price).quantize(Decimal("0.01"))
            total_quantity += quantity
            total_value += row.stock_value
        return total_quantity, total_value

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        total_quantity, total_value = self._decorate(context["records"])
        context["page_total_quantity"] = total_quantity
        context["page_total_value"] = total_value
        is_service = self.item_kind() == INVENTORY_KIND_SERVICE
        context["active_tab"] = "services" if is_service else "products"
        context["is_service_tab"] = is_service
        context["product_value"] = INVENTORY_KIND_PRODUCT
        context["service_value"] = INVENTORY_KIND_SERVICE
        return context


class ItemListView(ItemStockListMixin, ListView):
    template_name = "inventory/item_list.html"

    def get_context_data(self, **kwargs):
        # Items are the subsidiary ledger behind the balance sheet's Inventory
        # account, so the page states whether the two currently agree.
        from apps.finance.services import inventory_control_summary  # lazy: finance imports inventory

        context = super().get_context_data(**kwargs)
        context["title"] = "Services" if context["is_service_tab"] else "Inventory Items"
        context["create_url"] = reverse_lazy("inventory:item_create")
        # A service holds no stock, so the Inventory control banner belongs only
        # on the products tab, where the figures it reconciles are shown.
        context["control_account"] = None if context["is_service_tab"] else inventory_control_summary()
        return context


class ItemPrintView(PrintContextMixin, ItemStockListMixin, ListView):
    """Every item under the filters currently applied, unpaginated."""

    action = "index"
    template_name = "inventory/item_print.html"
    paginate_by = None

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["print_back_url"] = reverse_lazy("inventory:item_list")
        return context


class ItemExportView(ItemStockListMixin, ListView):
    """The item list as a spreadsheet, under the filters currently applied.

    Written as UTF-8 CSV rather than a real workbook: Excel opens it directly
    and it keeps the export dependency-free, matching the voucher export.
    """

    action = "index"
    paginate_by = None

    def get(self, request, *args, **kwargs):
        rows = list(self.get_queryset())
        total_quantity, total_value = self._decorate(rows)
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        stamp = timezone.localdate().isoformat()
        response["Content-Disposition"] = f'attachment; filename="inventory-items-{stamp}.csv"'
        # Excel reads a UTF-8 CSV correctly only when it starts with the BOM.
        response.write("﻿")
        writer = csv.writer(response)
        writer.writerow(["Item Name", "Code", "Category", "UOM", "Stock Qty", "Current Price", "Last Price", "Stock Value", "Status"])
        for row in rows:
            stock = getattr(row, "stock", None)
            writer.writerow([
                row.item_name,
                row.code,
                row.item_class.title if row.item_class_id else "",
                uom_title(row),
                stock.current_quantity if stock else "",
                stock.current_price if stock else "",
                stock.last_price if stock else "",
                row.stock_value,
                row.get_status_display(),
            ])
        writer.writerow(["Totals", "", "", "", total_quantity, "", "", total_value, ""])
        return response


class ItemImportView(PagePermissionRequiredMixin, FormView):
    """Create items in bulk from a CSV.

    Every row is validated through ``InventoryItemForm`` so an import cannot
    write anything the Add Item screen would have rejected, and the whole file
    is applied in one transaction: a partly-imported list is worse than none.
    """

    page = "inventory.items"
    action = "add"
    form_class = InventoryItemImportForm
    template_name = "inventory/item_import.html"
    success_url = reverse_lazy("inventory:item_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Import Items"
        context["sample_url"] = reverse_lazy("inventory:item_import_sample")
        return context

    @staticmethod
    def _lookup(model, value, fields):
        """Find a master record by any of ``fields``, case-insensitively."""
        value = (value or "").strip()
        if not value:
            return None
        for field in fields:
            match = model.objects.filter(**{f"{field}__iexact": value}).first()
            if match:
                return match
        return None

    @staticmethod
    def _read_csv(upload):
        """(header names, row dicts) from a CSV upload."""
        # utf-8-sig: a CSV saved by Excel carries a BOM, which would otherwise
        # become part of the first header name.
        text = upload.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        headers = [(name or "").strip().lower() for name in (reader.fieldnames or [])]
        rows = [{(k or "").strip().lower(): str(v or "").strip() for k, v in raw.items() if k} for raw in reader]
        return headers, rows

    @staticmethod
    def _read_xlsx(upload):
        """(header names, row dicts) from the first sheet of an .xlsx upload."""
        from openpyxl import load_workbook

        # data_only: a sheet built with formulas hands over the cached results
        # rather than "=A1*2". read_only keeps a long sheet off the heap.
        workbook = load_workbook(upload, data_only=True, read_only=True)
        try:
            sheet = workbook.worksheets[0]
            grid = sheet.iter_rows(values_only=True)
            try:
                header_row = next(grid)
            except StopIteration:
                return [], []
            headers = [str(name or "").strip().lower() for name in header_row]
            rows = []
            for values in grid:
                row = {}
                for index, name in enumerate(headers):
                    if not name:
                        continue
                    value = values[index] if index < len(values) else None
                    # A price typed as a number arrives as 100.0; str() of that
                    # is still what the item form parses, so no rounding here.
                    row[name] = "" if value is None else str(value).strip()
                rows.append(row)
            return headers, rows
        finally:
            workbook.close()

    def form_valid(self, form):
        upload = form.cleaned_data["file"]
        update_existing = form.cleaned_data["update_existing"]
        try:
            if upload.name.lower().endswith(".xlsx"):
                headers, data_rows = self._read_xlsx(upload)
            else:
                headers, data_rows = self._read_csv(upload)
        except UnicodeDecodeError:
            form.add_error("file", "File is not valid UTF-8 text. Re-save it as CSV UTF-8, or upload the .xlsx instead.")
            return self.form_invalid(form)
        except Exception:
            # Anything openpyxl throws on a corrupt or mislabelled workbook.
            form.add_error("file", "File could not be read. Check it is a real .xlsx workbook or a plain CSV.")
            return self.form_invalid(form)

        missing = [name for name in ("item_name", "item_class", "uom") if name not in headers]
        if missing:
            form.add_error("file", f"Missing required column(s): {', '.join(missing)}.")
            return self.form_invalid(form)

        created = updated = 0
        errors = []
        try:
            with transaction.atomic():
                for line_no, row in enumerate(data_rows, start=2):  # row 1 is the header
                    if not any(row.values()):
                        continue  # trailing blank line

                    item_class = self._lookup(InventoryClass, row.get("item_class"), ("title", "class_code"))
                    uom = self._lookup(UOM, row.get("uom"), ("title", "code"))
                    if not item_class:
                        errors.append(f"Row {line_no}: unknown item class '{row.get('item_class', '')}'.")
                        continue
                    # A blank unit column is allowed, an unreadable one is not:
                    # silently dropping a typo would file the item unmeasured.
                    if not uom and (row.get("uom") or "").strip():
                        errors.append(f"Row {line_no}: unknown UOM '{row.get('uom', '')}'.")
                        continue

                    existing = InventoryItem.objects.filter(item_name__iexact=row.get("item_name", "")).first()
                    if existing and not update_existing:
                        errors.append(f"Row {line_no}: '{existing.item_name}' already exists.")
                        continue

                    data = {
                        "item_name": row.get("item_name", ""),
                        # Left blank on purpose: the model derives the code from
                        # the class prefix, so imports match hand-added items.
                        "code": existing.code if existing else "",
                        "category": item_class.title,
                        "uom": uom.pk,
                        "item_bar_code": row.get("item_bar_code", ""),
                        "price": row.get("price") or "0",
                        "purchase_price": row.get("purchase_price") or "0",
                        # Bulk-loaded rows are stocked goods; a service is added
                        # one at a time from the Add Item screen.
                        "item_kind": INVENTORY_KIND_PRODUCT,
                        "status": STATUS_ACTIVE,
                        "imported": "L",
                        "inventory": "I",
                    }
                    item_form = InventoryItemForm(data, instance=existing)
                    if not item_form.is_valid():
                        detail = "; ".join(f"{field}: {msg[0]}" for field, msg in item_form.errors.items())
                        errors.append(f"Row {line_no}: {detail}")
                        continue

                    item = item_form.save(commit=False)
                    if not existing:
                        item.created_by = self.request.user
                    item.updated_by = self.request.user
                    item.save()
                    if existing:
                        updated += 1
                    else:
                        created += 1

                if errors:
                    # Nothing is kept when any row failed, so the file can be
                    # corrected and re-uploaded without hunting for duplicates.
                    raise _ImportRowError
        except _ImportRowError:
            context = self.get_context_data(form=form)
            context["row_errors"] = errors
            return self.render_to_response(context)

        messages.success(self.request, f"{created} item(s) created, {updated} updated.")
        return super().form_valid(form)


class _ImportRowError(Exception):
    """Internal: rolls the import transaction back when any row is rejected."""


class ItemImportSampleView(PagePermissionRequiredMixin, View):
    """A one-row template in the shape the importer expects.

    Defaults to .xlsx, since that is what most item lists are kept in;
    ``?format=csv`` hands back the same row as CSV.
    """

    page = "inventory.items"
    action = "add"

    @staticmethod
    def _sample_row():
        item_class = InventoryClass.objects.order_by("title").first()
        uom = UOM.objects.order_by("title").first()
        return [
            "Example Item",
            item_class.title if item_class else "Raw Material",
            uom.title if uom else "Each",
            "100.00",
            "80.00",
            "8901234567890",
        ]

    def get(self, request, *args, **kwargs):
        columns = list(InventoryItemImportForm.COLUMNS)
        if request.GET.get("format") == "csv":
            response = HttpResponse(content_type="text/csv; charset=utf-8")
            response["Content-Disposition"] = 'attachment; filename="item-import-sample.csv"'
            # Excel reads a UTF-8 CSV correctly only when it starts with the BOM.
            response.write("﻿")
            writer = csv.writer(response)
            writer.writerow(columns)
            writer.writerow(self._sample_row())
            return response

        from openpyxl import Workbook
        from openpyxl.styles import Font

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Items"
        sheet.append(columns)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.append(self._sample_row())
        for index, name in enumerate(columns, start=1):
            sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = max(len(name) + 4, 16)

        buffer = io.BytesIO()
        workbook.save(buffer)
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="item-import-sample.xlsx"'
        return response


class ItemToggleStatusView(InventoryManageMixin, View):
    page = "inventory.items"
    action = "edit"
    def post(self, request, pk):
        record = get_object_or_404(InventoryItem, pk=pk)
        record.status = STATUS_INACTIVE if record.status == STATUS_ACTIVE else STATUS_ACTIVE
        record.updated_by = request.user
        record.save(update_fields=["status", "updated_by", "updated_at"])
        return redirect(request.META.get("HTTP_REFERER") or reverse_lazy("inventory:item_list"))


class ItemNextCodeView(InventoryManageMixin, View):
    """The code the item form would assign under a typed category.

    A preview only: the number is settled again on save, so two operators
    filling the form at once still end up with different codes.
    """

    page = "inventory.items"

    def get(self, request, *args, **kwargs):
        title = (request.GET.get("category") or "").strip()
        item_class = InventoryClass.objects.filter(title__iexact=title).first() if title else None
        if item_class:
            prefix = (item_class.class_code or "ITM").upper()
        elif title:
            prefix = InventoryItemForm._next_class_code(title)
        else:
            prefix = "ITM"
        return JsonResponse({"prefix": prefix, "code": InventoryItem.next_code(prefix)})


class DirectPurchaseCreateView(InventoryManageMixin, View):
    """Enter a supplier bill without raising an order first.

    The screen is one form: the party at the top, the goods in the middle, the
    money at the bottom. Everything it posts goes through the same service the
    ordered route uses, so there is one set of books either way.
    """

    page = "inventory.purchase_orders"
    action = "add"
    template_name = "inventory/direct_purchase_form.html"

    def _context(self, **extra):
        items = (
            InventoryItem.objects
            .select_related("uom", "secondary_uom", "stock", "conversion__uom_from", "conversion__uom_to")
            .filter(status=STATUS_ACTIVE)
            .order_by("item_name")
        )
        context = {
            "title": "Purchase Invoice",
            "next_invoice_no": next_direct_purchase_number(),
            "suppliers": Supplier.objects.filter(status=STATUS_ACTIVE).order_by("name"),
            "units": UOM.objects.order_by("title"),
            "today": timezone.localdate(),
            "items_json": json.dumps([
                {
                    "id": item.pk,
                    "name": item.item_name,
                    "code": item.code,
                    "uom": item.uom_id or "",
                    "rate": float(item.purchase_price or 0),
                    # What is on the shelf now, so the line can show what this
                    # bill takes it to. A service is never stocked.
                    "stock": float(getattr(item.stock, "current_quantity", 0) or 0),
                    "stocked": item.item_kind == INVENTORY_KIND_PRODUCT,
                    "unit": uom_title(item),
                    "units": item_unit_options(item),
                }
                for item in items
            ]),
        }
        context.update(extra)
        return context

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name, self._context())

    def post(self, request, *args, **kwargs):
        posted = request.POST
        supplier_id = (posted.get("supplier") or "").strip()
        supplier = Supplier.objects.filter(pk=supplier_id).first() if supplier_id.isdigit() else None

        # Money boxes are grouped with commas on screen. They are stripped before
        # the form posts, but a figure that arrives grouped anyway must still be
        # read as the number it is rather than rejected.
        def decimal_of(raw, default="0"):
            text = (raw or "").strip().replace(",", "") or default
            return Decimal(text)

        def money(name, default="0"):
            try:
                return decimal_of(posted.get(name), default)
            except (InvalidOperation, ValueError):
                raise ValidationError(f"{name.replace('_', ' ').title()} must be a number.")

        # The rows arrive as parallel lists, so a blank row simply drops out.
        lines = []
        item_ids = posted.getlist("item_id")
        quantities = posted.getlist("quantity")
        rates = posted.getlist("rate")
        uom_ids = posted.getlist("line_uom")
        for index, raw_id in enumerate(item_ids):
            if not (raw_id or "").strip().isdigit():
                continue
            item = InventoryItem.objects.filter(pk=raw_id).first()
            if not item:
                continue
            try:
                quantity = decimal_of(quantities[index] if index < len(quantities) else "")
                rate = decimal_of(rates[index] if index < len(rates) else "")
            except (InvalidOperation, ValueError):
                # A line naming an item is never dropped in silence: the operator
                # meant to buy it, so the figure gets corrected rather than lost.
                messages.error(request, f"Check the quantity and price on the {item.item_name} line.")
                return render(request, self.template_name, self._context(posted=posted))
            if quantity > 0:
                # The unit the line was written in; the service restates it in
                # the item's own unit before anything is booked.
                raw_uom = (uom_ids[index] if index < len(uom_ids) else "") or ""
                uom = UOM.objects.filter(pk=raw_uom).first() if raw_uom.strip().isdigit() else None
                lines.append({"inventory_item": item, "quantity": quantity, "rate": rate, "uom": uom})

        try:
            bill_date = posted.get("bill_date") or str(timezone.localdate())
            order, net = create_direct_purchase(
                supplier=supplier,
                bill_number=(posted.get("bill_number") or "").strip(),
                bill_date=bill_date,
                lines=lines,
                discount_amount=money("discount_amount"),
                tax_amount=money("tax_amount"),
                paid_amount=money("paid_amount"),
                remarks=(posted.get("remarks") or "").strip(),
                user=request.user,
            )
        except ValidationError as error:
            for message in error.messages:
                messages.error(request, message)
            return render(request, self.template_name, self._context(posted=posted))

        messages.success(request, f"Purchase {order.purchase_num} saved for {net}.")
        if "save_and_new" in posted:
            return redirect("inventory:direct_purchase_create")
        # Straight to the printable copy of the bill just saved, rather than the
        # record screen the operator would have to hunt the print link on.
        if "save_and_print" in posted:
            return redirect("inventory:purchase_order_print", pk=order.pk)
        return redirect("inventory:purchase_order_detail", pk=order.pk)


class ItemConversionOptionsView(InventoryManageMixin, View):
    """The rates already on file between two units.

    The item form offers these as a pick list rather than asking the operator
    to retype a figure the units screen already holds.
    """

    page = "inventory.items"
    action = "add"

    def get(self, request, *args, **kwargs):
        base = request.GET.get("base") or ""
        secondary = request.GET.get("secondary") or ""
        if not base or not secondary:
            return JsonResponse({"options": []})

        rows = UOMConversion.objects.filter(uom_from_id=base, uom_to_id=secondary).select_related("uom_from", "uom_to")
        return JsonResponse({
            "options": [
                {
                    "id": row.pk,
                    "factor": f"{row.conversion_factor.normalize():f}",
                    "base": row.uom_from.title,
                    "secondary": row.uom_to.title,
                }
                for row in rows
            ]
        })


class ItemCreateView(EmbeddedCreateMixin, InventoryManageMixin, CreateView):
    page = "inventory.items"
    model = InventoryItem
    form_class = InventoryItemForm
    template_name = "inventory/item_form.html"
    success_url = reverse_lazy("inventory:item_list")
    success_message = "Item saved."
    embed_message_type = "item:saved"
    # Everything the first screen does not ask for; the Other tab renders these
    # by name so the layout never silently drops a field the form still posts.
    # `conversion` is not here: the unit dialog owns it now.
    other_fields = ("item_bar_code", "imported", "inventory", "status")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Inventory Item"
        context["other_fields"] = list(self.other_fields)
        context["product_value"] = INVENTORY_KIND_PRODUCT
        context["service_value"] = INVENTORY_KIND_SERVICE
        return context

    def embed_payload(self, obj):
        # What a line fills in from a pick: unit, expected cost and the stock
        # standing behind it, which an opening quantity may already have moved.
        stock = getattr(obj, "stock", None)
        return {
            "id": obj.pk,
            "name": obj.item_name,
            "code": obj.code,
            "uom": obj.uom_id or "",
            "rate": float(obj.purchase_price or 0),
            "stock": float(getattr(stock, "current_quantity", 0) or 0),
            "stocked": obj.item_kind == INVENTORY_KIND_PRODUCT,
            "unit": uom_title(obj),
            "units": item_unit_options(obj),
        }

    def form_valid(self, form):
        response = super().form_valid(form)
        quantity = form.cleaned_data.get("opening_quantity")
        if quantity:
            set_opening_stock(
                inventory_item=self.object,
                quantity=quantity,
                price=form.cleaned_data.get("opening_price"),
                opening_date=form.cleaned_data.get("opening_date"),
                user=self.request.user,
            )
        if self.is_embedded():
            return self.embed_saved_response()
        return response

    def get_success_url(self):
        # "Save & New" keeps the operator on a blank form for the next item.
        if "save_and_new" in self.request.POST:
            return reverse_lazy("inventory:item_create")
        # Otherwise stay on the record just saved, so it can be checked or
        # corrected without hunting for it in the list.
        return reverse_lazy("inventory:item_update", kwargs={"pk": self.object.pk})


class ItemUpdateView(ItemCreateView, UpdateView):
    success_message = "Item updated."


class LedgerListView(InventoryListMixin, ListView):
    page = "inventory.item_ledger"
    template_name = "inventory/ledger_list.html"
    context_object_name = "ledgers"
    queryset = ItemLedger.objects.select_related("inventory_item").order_by("-transaction_date", "-id")
    search_fields = ("transaction_id", "transaction_no", "item_code", "item_name", "ref_no", "transaction_type", "transaction_date", "old_quantity", "quantity", "new_quantity")
    filter_fields = {"item": "inventory_item_id", "type": "transaction_type"}
    date_filters = [{"field": "transaction_date", "label": "Transaction date"}]

    def get_filter_specs(self):
        item_choices = list(InventoryItem.objects.order_by("item_name").values_list("id", "item_name"))
        return [
            {"name": "type", "label": "All types", "choices": INV_TRANSACTION_TYPE_CHOICES, "value": self.request.GET.get("type", "")},
            {"name": "item", "label": "All items", "choices": item_choices, "value": self.request.GET.get("item", "")},
        ]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        rows = list(context["ledgers"])

        receive_ids = [r.ref_id for r in rows if r.ref_table == "inv_purchase_order_item_received"]
        sale_ids = [r.ref_id for r in rows if r.ref_table == "inv_pos_details"]
        receive_map = {
            rec.pk: rec.purchase_order_item.purchase_order_id
            for rec in PurchaseOrderItemReceived.objects.filter(pk__in=receive_ids).select_related("purchase_order_item")
        }
        sale_map = dict(POSDetail.objects.filter(pk__in=sale_ids).values_list("pk", "pos_master_id"))

        for row in rows:
            url = None
            if row.ref_table == "inv_manual_transaction" and row.transaction_id:
                url = reverse_lazy("inventory:manual_transaction_print", kwargs={"tx_id": row.transaction_id})
            elif row.ref_table == "inv_purchase_order_item_received" and row.ref_id in receive_map:
                url = f"{reverse_lazy('inventory:grn_print', kwargs={'pk': receive_map[row.ref_id]})}?receipts={row.ref_id}&reprint=1"
            elif row.ref_table == "inv_pos_details" and row.ref_id in sale_map:
                url = reverse_lazy("inventory:pos_receipt", kwargs={"pk": sale_map[row.ref_id]})
            row.ref_url = url
        return context


class CustomerLedgerListView(InventoryListMixin, ListView):
    page = "inventory.customer_ledger"
    template_name = "inventory/customer_ledger_list.html"
    context_object_name = "ledgers"
    queryset = CustomerLedger.objects.select_related("customer").order_by("-transaction_date", "-id")
    search_fields = ("transaction_no", "customer__customer_name", "customer__customer_code")
    filter_fields = {"customer": "customer_id"}
    date_filters = [{"field": "transaction_date", "label": "Transaction date"}]

    def get_filter_specs(self):
        customer_choices = list(Customer.objects.order_by("customer_name").values_list("id", "customer_name"))
        return [{"name": "customer", "label": "All customers", "choices": customer_choices, "value": self.request.GET.get("customer", "")}]


class LedgerPrintView(PrintContextMixin, InventoryListMixin, ListView):
    page = "inventory.item_ledger"
    action = "view"
    template_name = "inventory/ledger_print.html"
    context_object_name = "ledgers"
    queryset = ItemLedger.objects.select_related("inventory_item").order_by("-transaction_date", "-id")
    search_fields = ("transaction_id", "transaction_no", "item_code", "item_name", "ref_no")
    paginate_by = None

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["print_back_url"] = reverse_lazy("inventory:ledger_list")
        return context


class PurchaseOrderQuickCreateView(InventoryManageMixin, View):
    page = "inventory.purchase_orders"
    action = "add"
    def post(self, request):
        item_ids = request.POST.getlist("item_id")
        qtys = request.POST.getlist("qty")
        rates = request.POST.getlist("rate")
        discounts = request.POST.getlist("discount")

        lines = []
        for i, item_id in enumerate(item_ids):
            if not item_id:
                continue
            qty = Decimal(int(float(qtys[i] or "0")))
            if qty <= 0:
                continue
            lines.append((int(item_id), qty, Decimal(rates[i] or "0").quantize(Decimal("0.01")), Decimal(discounts[i] or "0").quantize(Decimal("0.01"))))

        form = PurchaseOrderForm(request.POST)
        if not form.is_valid():
            for errors in form.errors.values():
                for error in errors:
                    messages.error(request, error)
            return redirect("inventory:purchase_order_list")

        if not lines:
            messages.error(request, "Add at least one item before saving.")
            return redirect("inventory:purchase_order_list")

        with transaction.atomic():
            order = form.save(commit=False)
            order.status = STATUS_DRAFT
            order.created_by = request.user
            order.updated_by = request.user
            order.save()
            for item_id, qty, rate, discount in lines:
                inv_item = InventoryItem.objects.get(pk=item_id)
                PurchaseOrderItem.objects.create(
                    purchase_order=order,
                    inventory_item=inv_item,
                    uom=inv_item.uom,
                    quantity=qty,
                    rate=rate,
                    discount_amount=discount,
                    created_by=request.user,
                    updated_by=request.user,
                )

        messages.success(request, f"Purchase order {order.purchase_num} created with {len(lines)} item(s).")
        return redirect(f"{reverse_lazy('inventory:purchase_order_list')}?open={order.pk}")


class PurchaseOrderDraftInitView(InventoryManageMixin, View):
    page = "inventory.purchase_orders"
    action = "add"
    """AJAX: create draft PO header, return pk."""
    def post(self, request):
        from django.http import JsonResponse
        form = PurchaseOrderForm(request.POST)
        if not form.is_valid():
            return JsonResponse({"error": str(form.errors)}, status=400)
        order = form.save(commit=False)
        order.status = STATUS_DRAFT
        order.created_by = request.user
        order.updated_by = request.user
        order.save()
        return JsonResponse({"pk": order.pk, "purchase_num": order.purchase_num})


class PurchaseOrderDraftFinalizeView(InventoryManageMixin, View):
    page = "inventory.purchase_orders"
    action = "add"
    """Add items to existing draft PO and raise it."""
    def post(self, request):
        draft_pk = request.POST.get("draft_pk")
        order = get_object_or_404(PurchaseOrder, pk=draft_pk, status=STATUS_DRAFT)

        item_ids = request.POST.getlist("item_id")
        qtys = request.POST.getlist("qty")
        rates = request.POST.getlist("rate")
        discounts = request.POST.getlist("discount")

        lines = []
        for i, item_id in enumerate(item_ids):
            if not item_id:
                continue
            qty = Decimal(int(float(qtys[i] or "0")))
            if qty <= 0:
                continue
            lines.append((int(item_id), qty, Decimal(rates[i] or "0").quantize(Decimal("0.01")), Decimal(discounts[i] or "0").quantize(Decimal("0.01"))))

        if not lines:
            messages.error(request, "Add at least one item before saving.")
            return redirect("inventory:purchase_order_list")

        with transaction.atomic():
            for item_id, qty, rate, discount in lines:
                inv_item = InventoryItem.objects.get(pk=item_id)
                PurchaseOrderItem.objects.create(
                    purchase_order=order,
                    inventory_item=inv_item,
                    uom=inv_item.uom,
                    quantity=qty,
                    rate=rate,
                    discount_amount=discount,
                    created_by=request.user,
                    updated_by=request.user,
                )
            order.status = STATUS_RAISED
            order.updated_by = request.user
            order.save(update_fields=["status", "updated_by", "updated_at"])

        messages.success(request, f"Purchase order {order.purchase_num} raised with {len(lines)} item(s).")
        return redirect("inventory:purchase_order_print", pk=order.pk)


class PurchaseOrderRaiseView(InventoryManageMixin, View):
    """Release a draft order to the supplier.

    Guarded by ``edit`` and not by ``approve``, because most orders are within
    the buyer's own limit and releasing those is ordinary work. The service
    decides whether this particular order needed a second signature, which is
    the only place that can be decided -- it depends on the amount.
    """

    page = "inventory.purchase_orders"
    action = "edit"

    def post(self, request, pk):
        order = get_object_or_404(PurchaseOrder, pk=pk)
        try:
            approve_purchase_order(order=order, user=request.user)
            messages.success(request, f"{order.purchase_num} approved and released to the supplier.")
        except ValidationError as exc:
            messages.error(request, "; ".join(exc.messages))
        return redirect(f"{reverse_lazy('inventory:purchase_order_list')}?open={order.pk}")


class PurchaseOrderCancelView(InventoryManageMixin, View):
    """Abandon an order nothing has arrived against.

    Not a delete: the number stays in the sequence and the reason stays on the
    record, so a cancelled order can be told apart from one that never existed.
    """

    page = "inventory.purchase_orders"
    action = "approve"

    def post(self, request, pk):
        order = get_object_or_404(PurchaseOrder, pk=pk)
        form = PurchaseOrderCancelForm(request.POST)
        if not form.is_valid():
            messages.error(request, "Pick a reason for cancelling this order.")
            return redirect("inventory:purchase_order_detail", pk=pk)
        try:
            cancel_purchase_order(order=order, reason=form.cleaned_data["reason"], user=request.user)
            messages.success(request, f"{order.purchase_num} cancelled. The number stays in the sequence.")
        except ValidationError as exc:
            messages.error(request, "; ".join(exc.messages))
        return redirect("inventory:purchase_order_detail", pk=pk)


class PurchaseOrderCloseShortView(InventoryManageMixin, View):
    """Give up on the balance of a part-delivered order.

    Creates no accounting entry -- an order never had one. What it releases is
    the commitment, so the outstanding quantity stops counting as goods on
    order and stops propping up a reorder decision that will never be met.
    """

    page = "inventory.purchase_orders"
    action = "approve"

    def post(self, request, pk):
        order = get_object_or_404(PurchaseOrder, pk=pk)
        form = PurchaseOrderCloseShortForm(request.POST)
        if not form.is_valid():
            messages.error(request, "Pick a reason for closing this order short.")
            return redirect("inventory:purchase_order_detail", pk=pk)
        try:
            closed = close_purchase_order_short(order=order, reason=form.cleaned_data["reason"], user=request.user)
            messages.success(
                request,
                f"{closed.purchase_num} closed short — {closed.short_qty} units "
                f"({closed.short_value}) released from what is on order.",
            )
        except ValidationError as exc:
            messages.error(request, "; ".join(exc.messages))
        return redirect("inventory:purchase_order_detail", pk=pk)


class PurchaseOrderReopenView(InventoryManageMixin, View):
    """Expect the balance again, because the goods turned up after all."""

    page = "inventory.purchase_orders"
    action = "approve"

    def post(self, request, pk):
        order = get_object_or_404(PurchaseOrder, pk=pk)
        try:
            reopen_purchase_order(order=order, user=request.user)
            messages.success(request, f"{order.purchase_num} re-opened — the balance is expected again.")
        except ValidationError as exc:
            messages.error(request, "; ".join(exc.messages))
        return redirect("inventory:purchase_order_detail", pk=pk)


class PurchaseApprovalLimitView(InventoryManageMixin, View):
    """Set what a buyer may commit without a second signature."""

    page = "inventory.purchase_orders"
    action = "approve"

    def post(self, request):
        form = PurchaseApprovalLimitForm(request.POST)
        if not form.is_valid():
            messages.error(request, "Enter a valid approval limit.")
        else:
            try:
                limit = set_purchase_order_approval_limit(form.cleaned_data["amount"], user=request.user)
                messages.success(request, f"Approval limit set to {limit}. It applies to orders raised from now on.")
            except ValidationError as exc:
                messages.error(request, "; ".join(exc.messages))
        return redirect("inventory:purchase_order_list")


class GoodsReceiptReverseView(InventoryManageMixin, View):
    """Withdraw a posted goods receipt by posting its mirror image.

    Held behind its own permission rather than behind ``edit``: whoever enters
    a receipt should not, by that fact alone, be able to make one disappear.
    """

    page = "inventory.grn"
    action = "reverse"

    def post(self, request, pk):
        receipt = get_object_or_404(PurchaseOrderItemReceived, pk=pk)
        form = ReversalReasonForm(request.POST)
        if not form.is_valid():
            messages.error(request, "A reversal needs a reason.")
            return redirect("inventory:grn_list")
        try:
            mirror = reverse_purchase_receipt(receipt=receipt, reason=form.cleaned_data["reason"], user=request.user)
            messages.success(
                request,
                f"{receipt.grn_number} reversed by {mirror.grn_number}. The original stays in the books "
                "with a nil net effect — nothing was deleted.",
            )
        except ValidationError as exc:
            messages.error(request, "; ".join(exc.messages))
        return redirect("inventory:grn_list")


class PurchaseInvoiceListView(InventoryListMixin, ListView):
    """Bills entered straight off the supplier's invoice, with no order first.

    Kept apart from the purchase orders screen on purpose: an order is a thing
    still owed, while these arrived and were received the moment they were
    entered, so the two lists answer different questions.
    """

    page = "inventory.purchase_orders"
    template_name = "inventory/purchase_invoice_list.html"
    context_object_name = "invoices"
    queryset = (
        PurchaseOrder.objects
        .filter(is_direct=True)
        .select_related("supplier")
        .prefetch_related("items")
        .order_by("-purchase_date", "-id")
    )
    search_fields = ("purchase_num", "supplier__name", "quot_num", "descr")
    filter_fields = {"supplier": "supplier_id"}
    date_filters = [{"field": "purchase_date", "label": "Invoice date"}]

    def get_filter_specs(self):
        supplier_choices = list(
            Supplier.objects.filter(status=STATUS_ACTIVE).order_by("name").values_list("id", "name")
        )
        return [
            {"name": "supplier", "label": "All suppliers", "choices": supplier_choices,
             "value": self.request.GET.get("supplier", "")},
        ]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # What each bill came to, from its own lines, so the figure on screen is
        # the one the books hold rather than a second total kept in step by hand.
        page_total = Decimal("0.00")
        for invoice in context["invoices"]:
            lines = list(invoice.items.all())
            invoice.line_count = len(lines)
            invoice.total_amount = sum((line.total_amount for line in lines), Decimal("0.00"))
            page_total += invoice.total_amount
        context["page_total"] = page_total
        context["invoice_count"] = context["paginator"].count if context.get("paginator") else len(context["invoices"])
        return context


class PurchaseOrderListView(SortableListMixin, InventoryListMixin, ListView):
    """Orders raised on suppliers: what is committed, and what it is waiting on.

    The screen is built around the question an order actually poses -- is it
    approved, has it arrived, has it been billed, is it late -- rather than
    around the row in the table. The tabs sort orders by which of those they are
    stuck on, and the tiles across the top count the same thing the rows below
    show, because both are read from one decorated set.
    """

    page = "inventory.purchase_orders"
    template_name = "inventory/purchase_order_list.html"
    context_object_name = "orders"
    paginate_by = 25
    queryset = (
        PurchaseOrder.objects
        .filter(is_direct=False)
        .select_related("supplier", "created_by")
        .prefetch_related("items__receipts", "items__uom", "bills")
        .order_by("-purchase_date", "-id")
    )
    search_fields = ("purchase_num", "supplier__name", "quot_num", "descr")
    filter_fields = {"supplier": "supplier_id"}
    date_filters = [{"field": "purchase_date", "label": "Order date"}]
    # Only the columns the database can order by. What has arrived and what is
    # billed are worked out per row after the query, so they cannot be sorted
    # on without pulling the whole table into memory -- their headings stay
    # plain rather than offering a sort that would quietly lie about the order.
    sort_fields = {
        "purchase_num": "seq_num",
        "purchase_date": ("purchase_date", "id"),
        "supplier": "supplier__name",
        "expected": "expected_date",
        "status": "status",
    }
    default_sort = "purchase_date"
    # Newest first: the order somebody raised this morning is the one they are
    # looking for, not the one from last quarter.
    default_sort_dir = "desc"

    PER_PAGE_OPTIONS = (10, 25, 50, 100)

    def current_tab(self):
        tab = self.request.GET.get("tab", TAB_ALL)
        return tab if tab in dict(TABS) else TAB_ALL

    def get_paginate_by(self, queryset):
        raw = (self.request.GET.get("per_page") or "").strip()
        if raw.isdigit() and int(raw) in self.PER_PAGE_OPTIONS:
            return int(raw)
        return self.paginate_by

    def filtered_queryset(self):
        """Everything the filter bar allows, before the tab narrows it.

        The tiles are counted over this: switching to one tab should not make
        the numbers above it change, because they are what the tabs are for.
        """
        return super().get_queryset().distinct()

    def get_queryset(self):
        queryset = self.filtered_queryset()
        tab = self.current_tab()
        if tab == TAB_UNBILLED:
            # Goods booked in under a receipt that carries no supplier invoice.
            queryset = queryset.filter(items__receipts__invoice_num="").distinct()
        statuses = TAB_STATUSES.get(tab)
        if statuses:
            queryset = queryset.filter(status__in=statuses)
        return queryset

    def get_filter_specs(self):
        supplier_choices = list(
            Supplier.objects.filter(status=STATUS_ACTIVE).order_by("name").values_list("id", "name")
        )
        return [
            {"name": "supplier", "label": "All suppliers", "choices": supplier_choices,
             "value": self.request.GET.get("supplier", "")},
        ]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        orders = list(context["orders"])
        decorate(orders)
        context["orders"] = orders
        # The rows on show add up to the figure under them, which is what
        # "Total shown" means -- not the whole filtered set.
        context["page_total"] = sum((order.total_amount for order in orders), Decimal("0.00"))
        context["order_count"] = context["paginator"].count if context.get("paginator") else len(orders)

        # Counted over everything the filters allow, so the tiles hold still as
        # the tabs are clicked through.
        everything = list(self.filtered_queryset())
        context["tiles"] = summarise(everything)
        def tab_count(key):
            if key == TAB_UNBILLED:
                return sum(1 for order in everything if order.unbilled_value)
            statuses = TAB_STATUSES.get(key)
            return sum(1 for order in everything
                       if not statuses or order.status in statuses)

        context["tabs"] = [
            {"key": key, "label": label, "on": key == self.current_tab(), "count": tab_count(key)}
            for key, label in TABS
        ]
        # Which columns this person chose to look at. A set, so the template
        # asks `{% if "billed" in columns %}` rather than walking a list per row.
        # What is actually on each order, for the approval dialog to show. Only
        # the orders still awaiting approval need it, so the page does not carry
        # a line list for rows that will never open the dialog.
        context["approval_lines"] = {
            order.pk: [
                {
                    "name": line.descr,
                    "qty": f"{line.quantity.normalize():f}" if line.quantity else "0",
                    "uom": line.uom.title if line.uom_id else "",
                    "rate": float(line.rate or 0),
                    "amount": float(line.total_amount or 0),
                }
                for line in order.items.all()
            ]
            for order in orders if order.status == STATUS_DRAFT
        }
        context["columns"] = visible_columns(self.request.session)
        context["column_menu"] = column_menu(self.request.session)
        # Where the total row puts its figure: under PO Value wherever that
        # column lands, with the label stretched to meet it. "actions" is drawn
        # as its own trailing cell rather than from the column set, so it is
        # counted once here and not twice.
        shown = [column.key for column in COLUMNS.columns if column.key in context["columns"]]
        span = len(shown) + 1
        context["column_span"] = span
        if "value" in shown:
            context["foot_lead_span"] = shown.index("value")
            context["foot_tail_span"] = span - shown.index("value") - 1
        else:
            context["foot_lead_span"] = span
            context["foot_tail_span"] = 0
        # The figure the approval gate is measured against, so the tile that
        # counts drafts can say why they are drafts rather than leaving it as
        # something only the person who set it up knows.
        context["approval_limit"] = purchase_order_approval_limit()
        # The figure the approval gate is measured against, so the tile that
        # counts drafts can say why they are drafts rather than leaving it as
        # something only the person who set it up knows.
        context["approval_limit"] = purchase_order_approval_limit()
        context["current_tab"] = self.current_tab()
        # The current view as a query string with the tab left out, so a tab
        # link only has to append its own and every other setting survives.
        carried = self.request.GET.copy()
        for key in ("tab", "page"):
            carried.pop(key, None)
        context["base_query"] = carried.urlencode()
        context["per_page"] = self.get_paginate_by(None)
        context["per_page_options"] = list(self.PER_PAGE_OPTIONS)
        # Whether anything is narrowing the list right now. Paging and column
        # choices are not filters, so they do not light the reset up.
        context["export_url"] = reverse_lazy("inventory:purchase_order_export")
        context["columns_url"] = reverse_lazy("inventory:purchase_order_columns")
        context["filters_active"] = any(
            (self.request.GET.get(key) or "").strip()
            for key in ("q", "supplier", "date_from", "date_to", "sort", "dir")
        ) or self.current_tab() != TAB_ALL
        return context


class PurchaseOrderExportView(InventoryListMixin, TableExportView):
    """The orders on screen, in whichever format was asked for.

    Same filters and same columns as the list whatever the format, so the file
    needs no explaining and the two can never drift apart.
    """

    page = "inventory.purchase_orders"
    columns = COLUMNS
    filename = "purchase-orders"
    title = "Purchase Orders"

    def get_rows(self):
        listing = PurchaseOrderListView(request=self.request, kwargs={}, args=())
        return decorate(list(listing.get_queryset()))


class PurchaseOrderColumnsView(InventoryListMixin, View):
    """Which columns this person wants on the purchase orders table.

    Kept in the session, not the database: a column choice is how one operator
    likes to look at the screen, and it should not change what anybody else
    sees. Only the index permission is needed, because choosing what to look at
    is not a change to anything.
    """

    page = "inventory.purchase_orders"

    def post(self, request, *args, **kwargs):
        set_visible_columns(request.session, request.POST.getlist("columns"))
        # Back to the view they were on. Built from the posted filters rather
        # than from the Referer, so nothing off this site can steer the redirect.
        carried = request.POST.get("back", "")
        query = urlencode([
            (key, value) for key, value in parse_qsl(carried, keep_blank_values=False)
            if key in ("q", "tab", "supplier", "date_from", "date_to", "per_page", "page")
        ])
        target = reverse_lazy("inventory:purchase_order_list")
        return redirect(f"{target}?{query}" if query else str(target))


class PurchaseReportView(InventoryListMixin, ListView):
    page = "inventory.purchase_report"
    template_name = "inventory/purchase_report.html"
    context_object_name = "orders"
    queryset = PurchaseOrder.objects.select_related("supplier").prefetch_related("items").order_by("-purchase_date", "-id")
    search_fields = ("purchase_num", "supplier__name", "quot_num")
    filter_fields = {"status": "status", "supplier": "supplier_id", "item": "items__inventory_item_id"}
    date_filters = [{"field": "purchase_date", "label": "Purchase date"}]

    def get_queryset(self):
        return super().get_queryset().distinct()

    def get_filter_specs(self):
        supplier_choices = list(Supplier.objects.filter(status=STATUS_ACTIVE).order_by("name").values_list("id", "name"))
        item_choices = list(InventoryItem.objects.filter(status=STATUS_ACTIVE).order_by("item_name").values_list("id", "item_name"))
        return [
            {"name": "status", "label": "All statuses", "choices": INV_PURCHASE_ORDER_STATUS_CHOICES, "value": self.request.GET.get("status", "")},
            {"name": "supplier", "label": "All suppliers", "choices": supplier_choices, "value": self.request.GET.get("supplier", "")},
            {"name": "item", "label": "All items", "choices": item_choices, "value": self.request.GET.get("item", "")},
        ]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        for order in context["orders"]:
            items = list(order.items.all())
            order.po_total = sum(i.total_amount for i in items)
        return context


class PurchaseOrderCreateView(InventoryManageMixin, View):
    """Raise an order on a supplier.

    Deliberately the same screen as the purchase invoice — party at the top,
    goods in the middle, money at the bottom — because it is the same entry an
    operator makes; the difference is only that nothing is received here, so
    there is no paid box and no stock impact until the goods arrive.
    """

    page = "inventory.purchase_orders"
    action = "add"
    template_name = "inventory/purchase_order_form.html"

    def _context(self, **extra):
        items = (
            InventoryItem.objects
            .select_related("uom", "secondary_uom", "stock", "conversion__uom_from", "conversion__uom_to")
            .filter(status=STATUS_ACTIVE)
            .order_by("item_name")
        )
        context = {
            "title": "Purchase Order",
            # What this site has taken off the form and what it has added.
            "layout": get_layout(),
            "extra_field_types": EXTRA_FIELD_TYPES,
            # A plain View builds its own context, so the permission the menu
            # is gated on has to be put there by hand.
            "can_edit": user_has_permission(self.request.user, f"{self.page}.edit"),
            "next_order_no": next_purchase_order_number(),
            "suppliers": Supplier.objects.filter(status=STATUS_ACTIVE).order_by("name"),
            "units": UOM.objects.order_by("title"),
            "today": timezone.localdate(),
            "items_json": json.dumps([
                {
                    "id": item.pk,
                    "name": item.item_name,
                    "code": item.code,
                    "uom": item.uom_id or "",
                    "rate": float(item.purchase_price or 0),
                    "stock": float(getattr(item.stock, "current_quantity", 0) or 0),
                    "stocked": item.item_kind == INVENTORY_KIND_PRODUCT,
                    "unit": uom_title(item),
                    "units": item_unit_options(item),
                }
                for item in items
            ]),
        }
        context.update(extra)
        return context

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name, self._context())

    def post(self, request, *args, **kwargs):
        posted = request.POST
        supplier_id = (posted.get("supplier") or "").strip()
        supplier = Supplier.objects.filter(pk=supplier_id).first() if supplier_id.isdigit() else None

        def decimal_of(raw, default="0"):
            text = (raw or "").strip().replace(",", "") or default
            return Decimal(text)

        def money(name, default="0"):
            try:
                return decimal_of(posted.get(name), default)
            except (InvalidOperation, ValueError):
                raise ValidationError(f"{name.replace('_', ' ').title()} must be a number.")

        lines = []
        item_ids = posted.getlist("item_id")
        quantities = posted.getlist("quantity")
        rates = posted.getlist("rate")
        uom_ids = posted.getlist("line_uom")
        for index, raw_id in enumerate(item_ids):
            if not (raw_id or "").strip().isdigit():
                continue
            item = InventoryItem.objects.filter(pk=raw_id).first()
            if not item:
                continue
            try:
                quantity = decimal_of(quantities[index] if index < len(quantities) else "")
                rate = decimal_of(rates[index] if index < len(rates) else "")
            except (InvalidOperation, ValueError):
                messages.error(request, f"Check the quantity and price on the {item.item_name} line.")
                return render(request, self.template_name, self._context(posted=posted))
            if quantity > 0:
                # With the unit column off nothing is posted, and the line is
                # taken as written in the item's own unit.
                raw_uom = (uom_ids[index] if index < len(uom_ids) else "") or ""
                uom = UOM.objects.filter(pk=raw_uom).first() if raw_uom.strip().isdigit() else None
                lines.append({"inventory_item": item, "quantity": quantity, "rate": rate, "uom": uom})

        # Whatever the site added to the form. A required one that was left
        # blank stops the save, the same as any other required box.
        extra_values, extra_error = read_extra_values(posted)
        if extra_error:
            messages.error(request, extra_error)
            return render(request, self.template_name, self._context(posted=posted))

        try:
            order, net = create_purchase_order(
                supplier=supplier,
                quot_num=(posted.get("quot_num") or "").strip(),
                quot_date=(posted.get("quot_date") or "") or None,
                order_date=posted.get("order_date") or str(timezone.localdate()),
                expected_date=(posted.get("expected_date") or "") or None,
                lines=lines,
                discount_amount=money("discount_amount"),
                tax_amount=money("tax_amount"),
                remarks=(posted.get("remarks") or "").strip(),
                extra_data=extra_values,
                # Saved as a draft until it is raised, unless the operator said
                # to raise it here — the list screen offers the same step.
                status=STATUS_RAISED if "save_and_raise" in posted else STATUS_DRAFT,
                user=request.user,
            )
        except ValidationError as error:
            for message in error.messages:
                messages.error(request, message)
            return render(request, self.template_name, self._context(posted=posted))

        messages.success(request, f"Purchase order {order.purchase_num} saved for {net}.")
        if "save_and_print" in posted:
            return redirect("inventory:purchase_order_print", pk=order.pk)
        # Orders are raised in runs, so saving one hands back an empty form
        # rather than the order just saved; the message names what was posted.
        return redirect("inventory:purchase_order_create")


class PurchaseOrderFormSettingsView(InventoryManageMixin, View):
    """The settings menu on the purchase order form: what it shows, and what it adds.

    Everything arrives as a plain POST and sends the operator back to the form,
    so the menu never has to keep a half-applied state of its own. Configuring
    the screen is an edit to how the site works, so it wants the manage
    permission and not merely the right to raise an order.
    """

    page = "inventory.purchase_orders"
    action = "edit"

    def post(self, request, *args, **kwargs):
        step = request.POST.get("step")

        if step == "fields":
            # The menu posts what stays on; anything not ticked comes off.
            shown = set(request.POST.getlist("shown"))
            set_hidden([field["code"] for field in get_layout()["optional_fields"]
                        if field["code"] not in shown])
            messages.success(request, "Form fields updated.")

        elif step == "add":
            error = add_extra_field(
                code=request.POST.get("code"),
                label=request.POST.get("label"),
                kind=request.POST.get("type"),
                required=request.POST.get("required") == "1",
                # One choice per line is how a list is typed; commas belong
                # inside a choice, not between them.
                options=(request.POST.get("options") or "").splitlines(),
            )
            if error:
                messages.error(request, error)
            else:
                messages.success(request, "Field added to the form.")

        elif step == "remove":
            remove_extra_field((request.POST.get("code") or "").strip())
            messages.success(request, "Field removed from the form. What earlier orders recorded under it is kept.")

        return redirect("inventory:purchase_order_create")


class PurchaseOrderUpdateView(InventoryManageMixin, View):
    page = "inventory.purchase_orders"
    action = "edit"
    def get(self, request, pk):
        return self._blocked(request, pk)

    def post(self, request, pk):
        return self._blocked(request, pk)

    def _blocked(self, request, pk):
        messages.error(request, "A purchase order cannot be edited once it is created.")
        return redirect("inventory:purchase_order_detail", pk=pk)


class PurchaseOrderDetailView(InventoryListMixin, DetailView):
    page = "inventory.purchase_orders"
    model = PurchaseOrder
    template_name = "inventory/purchase_order_detail.html"
    context_object_name = "order"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        item_form = PurchaseOrderItemForm()
        used_item_ids = self.object.items.values_list("inventory_item_id", flat=True)
        available_items = item_form.fields["inventory_item"].queryset.exclude(pk__in=used_item_ids).select_related("uom")
        item_form.fields["inventory_item"].queryset = available_items
        context["item_form"] = item_form
        context["item_uom_map"] = {str(i.pk): {"name": i.item_name, "uom": uom_title(i)} for i in available_items}
        receive_form = ReceivePOForm(initial={"purchase_order_item": self.object.items.first()})
        receive_form.fields["purchase_order_item"].queryset = self.object.items.all()
        context["receive_form"] = receive_form

        # How this order may be ended, which depends on whether anything has
        # arrived against it. Cancel is for an order nothing came against;
        # close-short gives up the balance of one that was part delivered. The
        # template is told which applies rather than working it out itself.
        lines = list(self.object.items.all())
        anything_received = any((line.total_receive_qty or Decimal("0")) > 0 for line in lines)
        outstanding = sum((line.open_receive_qty for line in lines), Decimal("0"))
        context["can_cancel"] = self.object.status in (STATUS_DRAFT, STATUS_RAISED) and not anything_received
        context["can_close_short"] = anything_received and outstanding > Decimal("0.0005")
        context["is_closed_early"] = self.object.status in (STATUS_CANCELLED, STATUS_CLOSED_SHORT)
        context["outstanding_qty"] = outstanding
        context["cancel_form"] = PurchaseOrderCancelForm()
        context["close_short_form"] = PurchaseOrderCloseShortForm()
        context["reversal_form"] = ReversalReasonForm()
        # Receipts on this order, with whether each may still be withdrawn.
        receipts = []
        for line in lines:
            for receipt in line.receipts.all():
                ok, why = can_reverse_receipt(receipt)
                receipt.can_reverse = ok
                receipt.cannot_reverse_because = why
                receipts.append(receipt)
        context["receipts"] = sorted(receipts, key=lambda r: (r.receive_date, r.pk), reverse=True)
        return context


def redirect_after_item(request, pk):
    """Redirect back to the originating page with the PO collapsible auto-opened."""
    ref = request.META.get("HTTP_REFERER") or str(reverse_lazy("inventory:purchase_order_list"))
    parts = urlsplit(ref)
    query = [(key, value) for key, value in parse_qsl(parts.query) if key != "open"]
    query.append(("open", str(pk)))
    return redirect(urlunsplit((parts.scheme, parts.netloc, parts.path, urlencode(query), "")))


class PurchaseOrderItemCreateView(InventoryManageMixin, View):
    page = "inventory.purchase_orders"
    action = "add"
    def post(self, request, pk):
        order = get_object_or_404(PurchaseOrder, pk=pk)
        if order.status == STATUS_FULLY_RECEIVED:
            messages.error(request, "Fully received purchase order cannot be updated.")
            return redirect("inventory:purchase_order_detail", pk=pk)
        form = PurchaseOrderItemForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.purchase_order = order
            item.uom = item.inventory_item.uom
            item.created_by = request.user
            item.updated_by = request.user
            if item.is_duplicate_in_order():
                messages.error(request, "This item is already added to this purchase order.")
            else:
                item.save()
                messages.success(request, "Purchase order item saved.")
        else:
            for errors in form.errors.values():
                for error in errors:
                    messages.error(request, error)
        return redirect_after_item(request, pk)


class PurchaseOrderItemUpdateView(InventoryManageMixin, UpdateView):
    page = "inventory.purchase_orders"
    model = PurchaseOrderItem
    form_class = PurchaseOrderItemForm
    template_name = "inventory/simple_form.html"

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        if self.object.purchase_order.status != STATUS_RAISED:
            messages.error(request, "Only items of a created purchase order can be edited.")
            return redirect("inventory:purchase_order_detail", pk=self.object.purchase_order_id)
        return super().dispatch(request, *args, **kwargs)

    def get_initial(self):
        return {"uom_title": uom_title(self.object)}

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = f"Edit Item - {self.object.purchase_num}"
        return context

    def form_valid(self, form):
        item = form.save(commit=False)
        item.uom = item.inventory_item.uom
        item.updated_by = self.request.user
        if item.is_duplicate_in_order():
            form.add_error("inventory_item", "This item is already added to this purchase order.")
            return self.form_invalid(form)
        item.save()
        messages.success(self.request, "Purchase order item updated.")
        list_url = str(reverse_lazy("inventory:purchase_order_list"))
        return redirect(f"{list_url}?open={item.purchase_order_id}")


class PurchaseOrderPrintView(PrintContextMixin, InventoryListMixin, DetailView):
    page = "inventory.purchase_orders"
    model = PurchaseOrder
    template_name = "inventory/purchase_order_print.html"
    context_object_name = "order"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        items = list(self.object.items.filter(status=YES))
        context["items"] = items
        context["total_qty"] = sum((i.quantity or Decimal("0")) for i in items)
        context["total_discount"] = sum((i.discount_amount or Decimal("0")) for i in items)
        grand_total = sum((i.total_amount for i in items), Decimal("0"))
        context["grand_total"] = grand_total
        context["amount_in_words"] = amount_in_words(grand_total)
        # Back to the list the document belongs to: a direct bill lives on the
        # purchase invoices screen, which has no order row to reopen.
        context["print_back_url"] = (
            reverse_lazy("inventory:purchase_invoice_list") if self.object.is_direct
            else f"{reverse_lazy('inventory:purchase_order_list')}?open={self.object.pk}"
        )
        return context


class PurchaseOrderItemToggleStatusView(InventoryManageMixin, View):
    page = "inventory.purchase_orders"
    action = "edit"
    def post(self, request, pk):
        item = get_object_or_404(PurchaseOrderItem, pk=pk)
        item.status = NO if item.status == YES else YES
        item.updated_by = request.user
        item.save()
        return redirect_after_item(request, item.purchase_order_id)


class PurchaseReceiveView(InventoryManageMixin, View):
    page = "inventory.purchase_orders"
    action = "edit"
    def post(self, request, pk):
        order = get_object_or_404(PurchaseOrder, pk=pk)
        form = ReceivePOForm(request.POST)
        form.fields["purchase_order_item"].queryset = order.items.all()
        if form.is_valid():
            try:
                receive_purchase_order_item(user=request.user, **form.cleaned_data)
                messages.success(request, "GRN posted and stock updated.")
            except ValidationError as exc:
                messages.error(request, exc)
        else:
            for errors in form.errors.values():
                for error in errors:
                    messages.error(request, error)
        return redirect("inventory:purchase_order_detail", pk=pk)


def _current_draft_tx_id():
    return ManualTransaction.objects.filter(status=STATUS_DRAFT).order_by("-id").values_list("transaction_id", flat=True).first()


class ManualTransactionView(InventoryManageMixin, View):
    page = "inventory.manual_transaction"
    action = "index"
    template_name = "inventory/manual_transaction.html"

    def get(self, request):
        from django.shortcuts import render

        draft_tx_id = _current_draft_tx_id()
        rows = ManualTransaction.objects.select_related("inventory_item").filter(transaction_id=draft_tx_id, status=STATUS_DRAFT) if draft_tx_id else ManualTransaction.objects.none()
        used_item_ids = list(rows.values_list("inventory_item_id", flat=True))
        items = InventoryItem.objects.select_related("uom", "stock").exclude(pk__in=used_item_ids).order_by("item_name")
        batch_descr = rows.values_list("descr", flat=True).first() or ""
        batch_supplier = rows.values_list("supplier_id", flat=True).first()
        form = ManualTransactionForm(initial={"descr": batch_descr, "qty": 1, "supplier": batch_supplier})
        form.fields["inventory_item"].queryset = items
        form.fields["supplier"].queryset = Supplier.objects.filter(status=STATUS_ACTIVE).order_by("name")
        # group posted transactions by transaction_id for history table
        from itertools import groupby as _groupby
        posted_qs = ManualTransaction.objects.filter(status=STATUS_POSTED).order_by("transaction_id", "id")
        history = {}
        for tx_id, grp in _groupby(posted_qs, key=lambda r: r.transaction_id):
            rows_list = list(grp)
            history[tx_id] = {
                "rows": rows_list,
                "date": rows_list[0].created_at,
                "descr": rows_list[0].descr,
                "total_qty": sum(r.qty for r in rows_list),
                "total_amount": sum(r.qty * r.price for r in rows_list),
            }
        context = {
            "title": "Manual Stock Transaction",
            "form": form,
            "rows": rows,
            "transaction_id": draft_tx_id,
            "item_price_map": {str(i.pk): {"price": str(getattr(i, "stock", None) and i.stock.current_price or i.price), "qty": str(getattr(i, "stock", None) and i.stock.current_quantity or 0), "uom": uom_title(i)} for i in items},
            "history": history,
        }
        return render(request, self.template_name, context)


class ManualTransactionAddView(InventoryManageMixin, View):
    page = "inventory.manual_transaction"
    action = "add"
    def post(self, request):
        form = ManualTransactionForm(request.POST)
        if form.is_valid():
            draft_tx_id = _current_draft_tx_id() or generate_transaction_id("ADJ", ManualTransaction)
            if draft_tx_id and ManualTransaction.objects.filter(transaction_id=draft_tx_id, status=STATUS_DRAFT, inventory_item=form.cleaned_data["inventory_item"]).exists():
                messages.error(request, "This item is already added to the current batch.")
                return redirect("inventory:manual_transaction")
            row = form.save(commit=False)
            row.transaction_id = draft_tx_id
            row.status = STATUS_DRAFT
            row.created_by = request.user
            row.updated_by = request.user
            row.save()
            batch = ManualTransaction.objects.filter(transaction_id=draft_tx_id, status=STATUS_DRAFT)
            batch.update(supplier=row.supplier)
            if row.descr:
                batch.update(descr=row.descr)
            messages.success(request, "Entry added to draft.")
        else:
            for errors in form.errors.values():
                for error in errors:
                    messages.error(request, error)
        return redirect("inventory:manual_transaction")


class ManualTransactionToggleView(InventoryManageMixin, View):
    page = "inventory.manual_transaction"
    action = "edit"
    def post(self, request, pk):
        row = get_object_or_404(ManualTransaction, pk=pk, status=STATUS_DRAFT)
        row.selected = NO if row.selected == YES else YES
        row.updated_by = request.user
        row.save(update_fields=["selected", "updated_by", "updated_at"])
        return redirect("inventory:manual_transaction")


class ManualTransactionDeleteView(InventoryManageMixin, View):
    page = "inventory.manual_transaction"
    action = "delete"
    def post(self, request, pk):
        row = get_object_or_404(ManualTransaction, pk=pk, status=STATUS_DRAFT)
        row.delete()
        messages.success(request, "Entry removed.")
        return redirect("inventory:manual_transaction")


class ManualTransactionSubmitView(InventoryManageMixin, View):
    page = "inventory.manual_transaction"
    action = "add"
    def post(self, request):
        draft_tx_id = _current_draft_tx_id()
        if not draft_tx_id:
            messages.error(request, "No draft entries to submit.")
            return redirect("inventory:manual_transaction")
        try:
            count = finalize_manual_transaction(transaction_id=draft_tx_id, user=request.user)
            messages.success(request, f"{count} entries posted to ledger and stock updated.")
            return redirect("inventory:manual_transaction_print", tx_id=draft_tx_id)
        except ValidationError as exc:
            messages.error(request, exc)
        return redirect("inventory:manual_transaction")


class ManualTransactionPrintView(InventoryManageMixin, PrintContextMixin, View):
    page = "inventory.manual_transaction"
    action = "view"
    template_name = "inventory/manual_transaction_print.html"

    def get(self, request, tx_id):
        from django.shortcuts import render
        rows = ManualTransaction.objects.filter(transaction_id=tx_id, status=STATUS_POSTED).select_related("inventory_item__uom", "supplier").order_by("id")
        if not rows.exists():
            messages.error(request, "Transaction not found.")
            return redirect("inventory:manual_transaction")
        grand_total = sum(r.qty * r.price for r in rows)
        context = self.get_print_context(request)
        context.update({
            "tx_id": tx_id,
            "rows": rows,
            "grand_total": grand_total,
            "amount_in_words": amount_in_words(grand_total),
            "tx_date": rows[0].created_at,
            "descr": rows[0].descr,
            "supplier": rows[0].supplier,
            "prepared_by": rows[0].created_by,
        })
        return render(request, self.template_name, context)


class CustomerListView(BaseSimpleListView):
    page = "inventory.customers"
    model = Customer
    queryset = Customer.objects.select_related("city").order_by("customer_name")
    search_fields = ("customer_name", "customer_code", "customer_cell_no")
    filter_fields = {"status": "status"}
    extra_context = {"title": "Customers", "create_url": reverse_lazy("inventory:customer_create"), "edit_url_name": "inventory:customer_update", "status_toggle_url_name": "inventory:customer_toggle_status", "default_toggle_url_name": "inventory:customer_toggle_default", "columns": [("Name", "customer_name"), ("Code", "customer_code"), ("Cell", "customer_cell_no"), ("Status", "status_toggle"), ("Default Customer", "default_toggle")]}

    def get_filter_specs(self):
        return [{"name": "status", "label": "All statuses", "choices": RECORD_STATUS_CHOICES, "value": self.request.GET.get("status", "")}] 


class CustomerToggleStatusView(InventoryManageMixin, View):
    page = "inventory.customers"
    action = "edit"
    def post(self, request, pk):
        customer = get_object_or_404(Customer, pk=pk)
        customer.status = STATUS_INACTIVE if customer.status == STATUS_ACTIVE else STATUS_ACTIVE
        customer.updated_by = request.user
        if customer.status == STATUS_INACTIVE and customer.is_default:
            customer.is_default = False
            customer.save(update_fields=["status", "is_default", "updated_by", "updated_at"])
        else:
            customer.save(update_fields=["status", "updated_by", "updated_at"])
        return redirect(request.META.get("HTTP_REFERER") or reverse_lazy("inventory:customer_list"))


class CustomerToggleDefaultView(InventoryManageMixin, View):
    page = "inventory.customers"
    action = "edit"
    def post(self, request, pk):
        customer = get_object_or_404(Customer, pk=pk)
        if not customer.is_default and customer.status != STATUS_ACTIVE:
            messages.error(request, "An inactive customer cannot be set as default.")
            return redirect(request.META.get("HTTP_REFERER") or reverse_lazy("inventory:customer_list"))
        customer.is_default = not customer.is_default
        customer.updated_by = request.user
        customer.save()
        return redirect(request.META.get("HTTP_REFERER") or reverse_lazy("inventory:customer_list"))


class CustomerCreateView(EmbeddedCreateMixin, InventoryManageMixin, CreateView):
    page = "inventory.customers"
    model = Customer
    form_class = CustomerForm
    template_name = "inventory/simple_form.html"
    success_url = reverse_lazy("inventory:customer_list")
    success_message = "Customer saved."
    embed_message_type = "customer:saved"
    extra_context = {"title": "Customer"}

    def embed_payload(self, obj):
        return {"id": obj.pk, "name": obj.customer_name}

    def form_valid(self, form):
        creating = self.object is None  # None on create, set on update
        response = super().form_valid(form)
        if creating:
            node = create_customer_receivable_account(customer=self.object, user=self.request.user)
            opening_balance = form.cleaned_data.get("opening_balance")
            if node and opening_balance:
                node.opening_balance = opening_balance
                node.save(update_fields=["opening_balance", "updated_at"])
        if self.is_embedded():
            return self.embed_saved_response()
        return response


class CustomerUpdateView(CustomerCreateView, UpdateView):
    success_message = "Customer updated."


class SaleInvoiceListView(InventoryListMixin, ListView):
    """Sales entered as invoices, the counterpart of the purchase invoice list."""

    page = "inventory.pos_sales"
    template_name = "inventory/sale_invoice_list.html"
    context_object_name = "invoices"
    queryset = (
        POSMaster.objects
        .select_related("customer")
        .prefetch_related("items")
        .order_by("-sale_date", "-id")
    )
    search_fields = ("sale_num", "customer__customer_name", "invoice_num", "remarks")
    filter_fields = {"customer": "customer_id"}
    date_filters = [{"field": "sale_date", "label": "Sale date"}]

    def get_filter_specs(self):
        customer_choices = list(
            Customer.objects.filter(status=STATUS_ACTIVE).order_by("customer_name").values_list("id", "customer_name")
        )
        return [
            {"name": "customer", "label": "All customers", "choices": customer_choices,
             "value": self.request.GET.get("customer", "")},
        ]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        page_total = Decimal("0.00")
        for invoice in context["invoices"]:
            invoice.line_count = invoice.items.count()
            page_total += invoice.net_amount or Decimal("0.00")
        context["page_total"] = page_total
        context["invoice_count"] = context["paginator"].count if context.get("paginator") else len(context["invoices"])
        return context


class SaleInvoiceCreateView(InventoryManageMixin, View):
    """A sale written up as an invoice, rather than rung through the POS screen.

    Same entry as the purchase invoice, the other way round: the customer at the
    top, the goods in the middle, the money at the bottom. It posts through the
    same service the POS screen uses, so there is one set of books either way.
    """

    page = "inventory.pos_sales"
    action = "add"
    template_name = "inventory/sale_invoice_form.html"

    def _context(self, **extra):
        items = (
            InventoryItem.objects
            .select_related("uom", "secondary_uom", "stock", "conversion__uom_from", "conversion__uom_to")
            .filter(status=STATUS_ACTIVE)
            .order_by("item_name")
        )
        context = {
            "title": "Sale Invoice",
            "next_invoice_no": next_sale_invoice_number(),
            "customers": Customer.objects.filter(status=STATUS_ACTIVE).order_by("customer_name"),
            "units": UOM.objects.order_by("title"),
            "today": timezone.localdate(),
            "items_json": json.dumps([
                {
                    "id": item.pk,
                    "name": item.item_name,
                    "code": item.code,
                    "uom": item.uom_id or "",
                    # A sale is priced off the sale price, not what it cost.
                    "rate": float(item.price or 0),
                    "stock": float(getattr(item.stock, "current_quantity", 0) or 0),
                    "stocked": item.item_kind == INVENTORY_KIND_PRODUCT,
                    "unit": uom_title(item),
                    "units": item_unit_options(item),
                }
                for item in items
            ]),
        }
        context.update(extra)
        return context

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name, self._context())

    def post(self, request, *args, **kwargs):
        posted = request.POST
        customer_id = (posted.get("customer") or "").strip()
        customer = Customer.objects.filter(pk=customer_id).first() if customer_id.isdigit() else None

        def decimal_of(raw, default="0"):
            text = (raw or "").strip().replace(",", "") or default
            return Decimal(text)

        def money(name, default="0"):
            try:
                return decimal_of(posted.get(name), default)
            except (InvalidOperation, ValueError):
                raise ValidationError(f"{name.replace('_', ' ').title()} must be a number.")

        lines = []
        item_ids = posted.getlist("item_id")
        quantities = posted.getlist("quantity")
        prices = posted.getlist("rate")
        uom_ids = posted.getlist("line_uom")
        for index, raw_id in enumerate(item_ids):
            if not (raw_id or "").strip().isdigit():
                continue
            item = InventoryItem.objects.filter(pk=raw_id).first()
            if not item:
                continue
            try:
                quantity = decimal_of(quantities[index] if index < len(quantities) else "")
                price = decimal_of(prices[index] if index < len(prices) else "")
            except (InvalidOperation, ValueError):
                messages.error(request, f"Check the quantity and price on the {item.item_name} line.")
                return render(request, self.template_name, self._context(posted=posted))
            if quantity > 0:
                raw_uom = (uom_ids[index] if index < len(uom_ids) else "") or ""
                uom = UOM.objects.filter(pk=raw_uom).first() if raw_uom.strip().isdigit() else None
                lines.append({"inventory_item": item, "quantity": quantity, "price": price, "uom": uom})

        try:
            sale_date = posted.get("sale_date") or str(timezone.localdate())
            sale, net = create_direct_sale(
                customer=customer,
                sale_date=sale_date,
                lines=lines,
                discount_amount=money("discount_amount"),
                tax_amount=money("tax_amount"),
                paid_amount=money("paid_amount"),
                remarks=(posted.get("remarks") or "").strip(),
                user=request.user,
            )
        except ValidationError as error:
            for message in error.messages:
                messages.error(request, message)
            return render(request, self.template_name, self._context(posted=posted))

        messages.success(request, f"Sale {sale.sale_num} saved for {net}.")
        if "save_and_print" in posted:
            return redirect("inventory:pos_receipt", pk=sale.pk)
        return redirect("inventory:pos_detail", pk=sale.pk)


class POSListView(InventoryManageMixin, View):
    page = "inventory.pos_sales"
    action = "index"
    template_name = "inventory/pos_list.html"

    def get(self, request):
        from django.shortcuts import render

        items = [
            {"id": s.inventory_item_id, "name": s.item_name, "price": float(s.current_price or 0), "stock": float(s.current_quantity or 0)}
            for s in Stock.objects.filter(status=STATUS_ACTIVE, current_quantity__gt=0).order_by("item_name")
        ]
        context = {
            "master_form": POSMasterForm(),
            "items_json": items,
            "recent_sales": POSMaster.objects.select_related("customer").filter(posted=YES).order_by("-id")[:10],
        }
        return render(request, self.template_name, context)


class POSCheckoutView(InventoryManageMixin, View):
    page = "inventory.pos_sales"
    action = "add"
    def post(self, request):
        item_ids = request.POST.getlist("item_id")
        qtys = request.POST.getlist("qty")
        prices = request.POST.getlist("price")
        discounts = request.POST.getlist("discount")

        lines = []
        for idx, item_id in enumerate(item_ids):
            if not item_id:
                continue
            qty = Decimal(qtys[idx] or "0").quantize(Decimal("0.01"))
            if qty <= 0:
                continue
            price = Decimal(prices[idx] or "0").quantize(Decimal("0.01"))
            discount = Decimal(discounts[idx] or "0").quantize(Decimal("0.01"))
            lines.append((int(item_id), qty, price, discount))

        if not lines:
            messages.error(request, "Add at least one item with quantity before posting.")
            return redirect("inventory:pos_list")

        needed = {}
        for item_id, qty, _price, _disc in lines:
            needed[item_id] = needed.get(item_id, Decimal("0")) + qty
        for stock in Stock.objects.filter(inventory_item_id__in=needed):
            if needed[stock.inventory_item_id] > (stock.current_quantity or Decimal("0")):
                messages.error(request, f"Insufficient stock for {stock.item_name} (available {stock.current_quantity}).")
                return redirect("inventory:pos_list")

        return self._checkout(request, lines)

    @transaction.atomic
    def _checkout(self, request, lines):
        sale = POSMaster.objects.create(
            transaction_id=generate_transaction_id("SAL", POSMaster),
            sale_date=request.POST.get("sale_date") or timezone.localdate(),
            pay_mode=request.POST.get("pay_mode") or "cash",
            customer_id=request.POST.get("customer") or None,
            remarks=f"Walking customer {timezone.now():%Y-%m-%d %H:%M:%S}",
            created_by=request.user,
            updated_by=request.user,
        )
        for item_id, qty, price, discount in lines:
            POSDetail.objects.create(
                pos_master=sale,
                inventory_item_id=item_id,
                quantity=qty,
                price=price,
                discount_amount=discount,
                created_by=request.user,
                updated_by=request.user,
            )
        post_sale(sale=sale, user=request.user)
        messages.success(request, f"Sale {sale.sale_num} posted — {len(lines)} item(s), stock updated.")
        return redirect("inventory:pos_receipt", pk=sale.pk)


class POSReceiptView(PrintContextMixin, InventoryListMixin, DetailView):
    page = "inventory.pos_sales"
    model = POSMaster
    template_name = "inventory/pos_receipt.html"
    context_object_name = "sale"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        items = self.object.items.all()
        context["items"] = items
        context["total_qty"] = sum((i.quantity or Decimal("0") for i in items), Decimal("0"))
        context["total_discount"] = sum((i.discount_amount or Decimal("0") for i in items), Decimal("0"))
        return context


class POSCreateView(InventoryManageMixin, CreateView):
    page = "inventory.pos_sales"
    model = POSMaster
    form_class = POSMasterForm
    template_name = "inventory/simple_form.html"
    success_url = reverse_lazy("inventory:pos_list")
    success_message = "Sale saved."
    extra_context = {"title": "POS Sale"}

    def form_valid(self, form):
        if not form.instance.transaction_id:
            form.instance.transaction_id = generate_transaction_id("SAL", POSMaster)
        if not form.instance.remarks:
            form.instance.remarks = f"Walking customer {timezone.now():%Y-%m-%d %H:%M:%S}"
        return super().form_valid(form)


class POSUpdateView(POSCreateView, UpdateView):
    success_message = "Sale updated."

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        if self.object.posted == YES:
            messages.error(request, "Posted sale cannot be updated.")
            return redirect("inventory:pos_detail", pk=self.object.pk)
        return super().dispatch(request, *args, **kwargs)


class POSDetailView(InventoryListMixin, DetailView):
    page = "inventory.pos_sales"
    model = POSMaster
    template_name = "inventory/pos_detail.html"
    context_object_name = "sale"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["item_form"] = POSDetailForm()
        items = self.object.items.all()
        bill_amount = sum((i.total_price or Decimal("0") for i in items), Decimal("0"))
        discount_total = sum((i.discount_amount or Decimal("0") for i in items), Decimal("0"))
        tax_total = sum((i.tax_amount or Decimal("0") for i in items), Decimal("0"))
        net_amount = bill_amount - discount_total + tax_total
        context["bill_amount"] = bill_amount
        context["discount_total"] = discount_total
        context["tax_total"] = tax_total
        context["net_amount"] = net_amount
        context["payable_amount"] = net_amount
        return context


class POSReturnQuickCreateView(InventoryManageMixin, View):
    page = "inventory.pos_returns"
    action = "add"
    def post(self, request):
        sale_id = request.POST.get("pos_master")
        if not sale_id:
            messages.error(request, "Select a sale first.")
            return redirect("inventory:pos_return_list")

        sale = get_object_or_404(POSMaster, pk=sale_id, posted=YES)
        detail_ids = request.POST.getlist("detail_id")
        return_qtys = request.POST.getlist("return_qty")

        lines = []
        for i, detail_id in enumerate(detail_ids):
            if not detail_id:
                continue
            qty = Decimal(return_qtys[i] or "0").quantize(Decimal("0.0001"))
            if qty <= 0:
                continue
            lines.append((int(detail_id), qty))

        if not lines:
            messages.error(request, "Select at least one item with return quantity.")
            return redirect("inventory:pos_return_list")

        pay_mode = request.POST.get("pay_mode") or "cash"
        adjusted_amount = Decimal(request.POST.get("adjusted_amount") or "0").quantize(Decimal("0.01"))
        return_date = request.POST.get("return_date") or timezone.localdate()

        with transaction.atomic():
            sale_return = POSReturnMaster.objects.create(
                transaction_id=generate_transaction_id("SRT", POSReturnMaster),
                pos_master=sale,
                return_date=return_date,
                pay_mode=pay_mode,
                adjusted_amount=adjusted_amount,
                created_by=request.user,
                updated_by=request.user,
            )
            for detail_id, qty in lines:
                pos_detail = get_object_or_404(POSDetail, pk=detail_id, pos_master=sale)
                if qty > pos_detail.quantity:
                    messages.error(request, f"{pos_detail.item_name}: return qty ({qty}) cannot exceed sale qty ({pos_detail.quantity}).")
                    return redirect("inventory:pos_return_list")
                POSReturnDetail.objects.create(
                    pos_return_master=sale_return,
                    pos_detail=pos_detail,
                    quantity=qty,
                    created_by=request.user,
                    updated_by=request.user,
                )

        try:
            post_sale_return(sale_return=sale_return, user=request.user)
        except ValidationError as exc:
            messages.error(request, str(exc))
            return redirect("inventory:pos_return_list")

        messages.success(request, f"Sale return {sale_return.return_num} posted — stock updated.")
        return redirect("inventory:pos_return_receipt", pk=sale_return.pk)


class POSReturnReceiptView(PrintContextMixin, InventoryListMixin, DetailView):
    page = "inventory.pos_returns"
    model = POSReturnMaster
    template_name = "inventory/pos_return_receipt.html"
    context_object_name = "sale_return"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        items = self.object.items.all()
        context["items"] = items
        context["total_qty"] = sum((i.quantity or Decimal("0") for i in items), Decimal("0"))
        context["total_return"] = sum((i.net_total or Decimal("0") for i in items), Decimal("0"))
        return context


class POSReturnListView(InventoryListMixin, ListView):
    page = "inventory.pos_returns"
    template_name = "inventory/pos_return_list.html"
    context_object_name = "returns"
    queryset = POSReturnMaster.objects.select_related("pos_master", "customer").filter(posted=YES).order_by("-return_date", "-id")
    search_fields = ("return_num", "transaction_id", "sale_num")
    date_filters = [{"field": "return_date", "label": "Return date"}]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        returned_sale_ids = POSReturnMaster.objects.filter(posted=YES).values_list("pos_master_id", flat=True)
        sales = POSMaster.objects.filter(posted=YES).exclude(pk__in=returned_sale_ids).prefetch_related("items__inventory_item").order_by("-sale_date", "-id")
        sales_json = []
        sale_items_json = {}
        for sale in sales:
            sales_json.append({
                "id": sale.pk,
                "sale_num": sale.sale_num,
                "customer": sale.customer.customer_name if sale.customer_id else "",
                "date": str(sale.sale_date),
                "net_amount": float(sale.net_amount),
            })
            sale_items_json[sale.pk] = [
                {
                    "id": item.pk,
                    "item_name": item.item_name,
                    "item_code": item.item_code,
                    "qty": float(item.quantity),
                    "price": float(item.price),
                    "net_total": float(item.net_total),
                }
                for item in sale.items.all()
            ]
        context["sales_json"] = sales_json
        context["sale_items_json"] = sale_items_json
        context["today"] = timezone.localdate()
        return context


class POSReturnCreateView(InventoryManageMixin, CreateView):
    page = "inventory.pos_returns"
    model = POSReturnMaster
    form_class = POSReturnMasterForm
    template_name = "inventory/simple_form.html"
    success_url = reverse_lazy("inventory:pos_return_list")
    success_message = "Sale return saved."
    extra_context = {"title": "POS Return"}

    def form_valid(self, form):
        if not form.instance.transaction_id:
            form.instance.transaction_id = generate_transaction_id("SRT", POSReturnMaster)
        return super().form_valid(form)


class POSReturnDetailView(InventoryListMixin, DetailView):
    page = "inventory.pos_returns"
    model = POSReturnMaster
    template_name = "inventory/pos_return_detail.html"
    context_object_name = "sale_return"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        form = POSReturnDetailForm()
        form.fields["pos_detail"].queryset = self.object.pos_master.items.all()
        context["item_form"] = form
        return context


class POSReturnItemCreateView(InventoryManageMixin, View):
    page = "inventory.pos_returns"
    action = "add"
    def post(self, request, pk):
        sale_return = get_object_or_404(POSReturnMaster, pk=pk)
        if sale_return.posted == YES:
            messages.error(request, "Posted sale return cannot be updated.")
            return redirect("inventory:pos_return_detail", pk=pk)
        form = POSReturnDetailForm(request.POST)
        form.fields["pos_detail"].queryset = sale_return.pos_master.items.all()
        if form.is_valid():
            item = form.save(commit=False)
            item.pos_return_master = sale_return
            item.created_by = request.user
            item.updated_by = request.user
            item.save()
            messages.success(request, "Return item saved.")
        else:
            for errors in form.errors.values():
                for error in errors:
                    messages.error(request, error)
        return redirect("inventory:pos_return_detail", pk=pk)


class POSReturnPostView(InventoryManageMixin, View):
    page = "inventory.pos_returns"
    action = "edit"
    def post(self, request, pk):
        record = get_object_or_404(POSReturnMaster, pk=pk)
        try:
            post_sale_return(sale_return=record, user=request.user)
            messages.success(request, "Sale return posted and stock updated.")
        except ValidationError as exc:
            messages.error(request, exc)
        return redirect("inventory:pos_return_detail", pk=pk)


class GRNListView(InventoryListMixin, ListView):
    page = "inventory.grn"
    template_name = "inventory/grn_list.html"
    context_object_name = "orders"
    queryset = PurchaseOrder.objects.select_related("supplier").prefetch_related("items__receipts", "bills").exclude(status__in=(STATUS_FULLY_RECEIVED, STATUS_CANCELLED, STATUS_CLOSED_SHORT)).order_by("-purchase_date", "-id")
    search_fields = ("purchase_num", "supplier__name", "quot_num")
    filter_fields = {"supplier": "supplier_id"}
    date_filters = [{"field": "purchase_date", "label": "Purchase date"}]

    PER_PAGE_OPTIONS = (10, 25, 50, 100)

    def get_paginate_by(self, queryset):
        raw = (self.request.GET.get("per_page") or "").strip()
        if raw.isdigit() and int(raw) in self.PER_PAGE_OPTIONS:
            return int(raw)
        return self.paginate_by

    def get_filter_specs(self):
        supplier_choices = list(Supplier.objects.filter(status=STATUS_ACTIVE).order_by("name").values_list("id", "name"))
        return [{"name": "supplier", "label": "All suppliers", "short_label": "Supplier",
                 "choices": supplier_choices, "value": self.request.GET.get("supplier", "")}]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # What the shared filter bar needs. No tabs, columns or export here --
        # the component leaves out whatever it is not given.
        carried = self.request.GET.copy()
        for key in ("tab", "page"):
            carried.pop(key, None)
        context["base_query"] = carried.urlencode()
        context["per_page"] = self.get_paginate_by(None)
        context["per_page_options"] = list(self.PER_PAGE_OPTIONS)
        context["filters_active"] = any(
            (self.request.GET.get(key) or "").strip()
            for key in ("q", "supplier", "date_from", "date_to")
        )
        context["columns"] = GRN_COLUMNS.visible(self.request.session)
        context["column_menu"] = GRN_COLUMNS.menu(self.request.session)
        context["columns_url"] = reverse_lazy("inventory:grn_columns")
        context["export_url"] = reverse_lazy("inventory:grn_export")
        # The expander handle and the GRN button are cells the table draws
        # itself rather than columns anyone may switch off.
        context["column_span"] = len(context["columns"]) + 2
        grns = list(
            PurchaseOrderItemReceived.objects
            .select_related("purchase_order_item__purchase_order", "inventory_item")
            .order_by("-receive_date", "-id")
        )
        # Whether each one may still be withdrawn, worked out here so the
        # template prints an answer rather than guessing at one.
        for grn in grns:
            ok, why = can_reverse_receipt(grn)
            grn.can_reverse = ok
            grn.cannot_reverse_because = why
        context["grns"] = grns
        context["reversal_form"] = ReversalReasonForm()
        for order in context["orders"]:
            items = list(order.items.all())
            order.po_total = sum(i.total_amount for i in items)
            order.ordered_total = Decimal("0")
            order.received_total = Decimal("0")
            order.balance_total = Decimal("0")
            for item in items:
                remaining = (item.quantity or Decimal("0")) - (item.total_receive_qty or Decimal("0"))
                item.remaining_qty = remaining if remaining > 0 else Decimal("0")

                # Every delivery against this line, oldest first, each carrying
                # what it took the running total to and what was still owed
                # afterwards. Stored ordering is newest-first, which is right
                # for a list and wrong for an account.
                ordered_qty = item.quantity or Decimal("0")
                running = Decimal("0")
                history = []
                for seq, receipt in enumerate(
                    sorted(item.receipts.all(), key=lambda r: (r.receive_date, r.pk)), start=1
                ):
                    taken = (receipt.quantity or Decimal("0")) + (receipt.extra_qty or Decimal("0"))
                    running += taken
                    balance = ordered_qty - running
                    history.append({
                        "seq": seq,
                        "receipt": receipt,
                        "quantity": taken,
                        "running": running,
                        "balance": balance if balance > 0 else Decimal("0"),
                    })
                item.history = history
                # What the line has actually taken in, read off the receipts
                # rather than off the running column the row also carries -- if
                # the two ever disagree, the receipts are the record.
                item.received_total = running
                order.ordered_total += ordered_qty
                order.received_total += running
                order.balance_total += item.remaining_qty
        return context


class GRNExportView(InventoryListMixin, TableExportView):
    """The goods receipt rows on screen, in whichever format was asked for."""

    page = "inventory.grn"
    columns = GRN_COLUMNS
    filename = "goods-receipts"
    title = "Goods Receipts"

    def get_rows(self):
        listing = GRNListView(request=self.request, kwargs={}, args=())
        orders = list(listing.get_queryset())
        # The two rolled-up figures the table shows are worked out on the way
        # to the page, so they are worked out here too rather than exporting
        # blanks for columns that are on screen.
        for order in orders:
            lines = list(order.items.all())
            order.po_total = sum((line.total_amount for line in lines), Decimal("0.00"))
            order.received_total = sum((line.total_receive_qty or Decimal("0") for line in lines), Decimal("0"))
            order.ordered_total = sum((line.quantity or Decimal("0") for line in lines), Decimal("0"))
            balance = order.ordered_total - order.received_total
            order.balance_total = balance if balance > 0 else Decimal("0")
        return orders


class GRNColumnsView(InventoryListMixin, View):
    """Which columns this person wants on the goods receipt table."""

    page = "inventory.grn"

    def post(self, request, *args, **kwargs):
        GRN_COLUMNS.choose(request.session, request.POST.getlist("columns"))
        carried = urlencode([
            (key, value) for key, value in parse_qsl(request.POST.get("back", ""), keep_blank_values=False)
            if key in ("q", "supplier", "date_from", "date_to", "per_page", "page")
        ])
        target = reverse_lazy("inventory:grn_list")
        return redirect(f"{target}?{carried}" if carried else str(target))


class GRNPrintView(PrintContextMixin, InventoryListMixin, DetailView):
    page = "inventory.grn"
    model = PurchaseOrder
    template_name = "inventory/grn_print.html"
    context_object_name = "order"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        receipt_pks_raw = self.request.GET.get("receipts", "")
        if receipt_pks_raw:
            pks = [p for p in receipt_pks_raw.split(",") if p.isdigit()]
            receipts = list(PurchaseOrderItemReceived.objects.filter(pk__in=pks).select_related("purchase_order_item"))
        else:
            receipts = []
            for item in self.object.items.prefetch_related("receipts").filter(status=YES):
                last = item.receipts.order_by("-id").first()
                if last:
                    receipts.append(last)
        for r in receipts:
            r.line_total = r.quantity * r.retail_price
        context["receipts"] = receipts
        context["total_qty"] = sum(r.quantity for r in receipts)
        context["grand_total"] = sum(r.line_total for r in receipts)
        context["po_total_qty"] = sum(r.purchase_order_item.quantity for r in receipts)
        context["receive_date"] = receipts[0].receive_date if receipts else timezone.localdate()
        context["prepared_by"] = self.request.user.get_full_name() or self.request.user.username
        context["is_reprint"] = self.request.GET.get("reprint") == "1"
        context["amount_in_words"] = amount_in_words(context["grand_total"])
        context["print_back_url"] = reverse_lazy("inventory:grn_list")
        return context


class GRNBulkReceiveView(InventoryManageMixin, View):
    page = "inventory.grn"
    action = "edit"
    def post(self, request, pk):
        order = get_object_or_404(PurchaseOrder, pk=pk)
        today = timezone.localdate()
        errors = []
        receipt_pks = []

        try:
            freight_total = Decimal(request.POST.get("freight_charges", "").strip() or "0")
        except Exception:
            freight_total = Decimal("0")
        if freight_total < 0:
            freight_total = Decimal("0")

        # First pass: collect valid receive lines and their cost value for freight allocation.
        recv_lines = []
        for item in order.items.filter(status=YES):
            raw = request.POST.get(f"recv_qty_{item.pk}", "").strip()
            if not raw:
                continue
            try:
                qty = Decimal(raw)
            except Exception:
                continue
            if qty <= 0:
                continue
            unit_cost = item.retail_price or item.rate or Decimal("0")
            recv_lines.append((item, qty, unit_cost, qty * unit_cost))

        total_value = sum((line[3] for line in recv_lines), Decimal("0"))

        try:
            with transaction.atomic():
                allocated = Decimal("0")
                for idx, (item, qty, unit_cost, line_value) in enumerate(recv_lines):
                    if freight_total <= 0 or total_value <= 0:
                        freight_amount = Decimal("0")
                    elif idx == len(recv_lines) - 1:
                        freight_amount = (freight_total - allocated).quantize(Decimal("0.01"))
                    else:
                        freight_amount = (freight_total * line_value / total_value).quantize(Decimal("0.01"))
                        allocated += freight_amount
                    try:
                        receipt = receive_purchase_order_item(
                            purchase_order_item=item,
                            quantity=qty,
                            extra_qty=Decimal("0"),
                            retail_price=unit_cost,
                            receive_date=today,
                            invoice_num="",
                            invoice_date=None,
                            rv_number="",
                            remarks="",
                            user=request.user,
                            freight_amount=freight_amount,
                        )
                        receipt_pks.append(str(receipt.pk))
                    except ValidationError as exc:
                        errors.append(str(exc))
                if errors:
                    raise ValidationError(errors)
        except ValidationError as exc:
            for msg in exc.messages:
                messages.error(request, msg)
            return redirect("inventory:grn_list")
        if receipt_pks:
            messages.success(request, f"{len(receipt_pks)} item(s) received for {order.purchase_num}.")
        return redirect(f"{reverse_lazy('inventory:grn_print', kwargs={'pk': pk})}?receipts={','.join(receipt_pks)}")


class PurchaseReturnListView(InventoryListMixin, ListView):
    page = "inventory.purchase_returns"
    template_name = "inventory/purchase_return_list.html"
    context_object_name = "returns"
    queryset = PurchaseReturnMaster.objects.filter(posted=YES).select_related("purchase_order", "supplier").order_by("-return_date", "-id")
    search_fields = ("return_num", "transaction_id", "purchase_order__purchase_num")
    filter_fields = {"supplier": "supplier_id"}
    date_filters = [{"field": "return_date", "label": "Return date"}]

    def get_filter_specs(self):
        supplier_choices = list(Supplier.objects.filter(status=STATUS_ACTIVE).order_by("name").values_list("id", "name"))
        return [{"name": "supplier", "label": "All suppliers", "choices": supplier_choices, "value": self.request.GET.get("supplier", "")}]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        all_orders = PurchaseOrder.objects.select_related("supplier").prefetch_related("items__inventory_item", "items__uom").filter(
            status__in=[STATUS_PARTIAL_RECEIVED, STATUS_FULLY_RECEIVED]
        ).order_by("-purchase_date", "-id")
        returnable_pks = []
        for o in all_orders:
            for item in o.items.all():
                if (item.total_receive_qty or Decimal("0")) <= 0:
                    continue
                already = PurchaseReturnDetail.objects.filter(
                    purchase_return_master__purchase_order=o,
                    purchase_return_master__posted=YES,
                    inventory_item=item.inventory_item,
                ).aggregate(t=Sum("quantity"))["t"] or Decimal("0")
                if item.total_receive_qty > already:
                    returnable_pks.append(o.pk)
                    break
        orders = all_orders.filter(pk__in=returnable_pks)
        pos_json = [
            {"id": o.pk, "purchase_num": o.purchase_num, "supplier": o.supplier.name, "date": str(o.purchase_date)}
            for o in orders
        ]
        items_json = {}
        for o in orders:
            rows = []
            for i in o.items.all():
                if (i.total_receive_qty or Decimal("0")) <= 0:
                    continue
                already = PurchaseReturnDetail.objects.filter(
                    purchase_return_master__purchase_order=o,
                    purchase_return_master__posted=YES,
                    inventory_item=i.inventory_item,
                ).aggregate(t=Sum("quantity"))["t"] or Decimal("0")
                returnable = i.total_receive_qty - already
                if returnable <= 0:
                    continue
                rows.append({
                    "poi_pk": i.pk,
                    "inventory_item_pk": i.inventory_item_id,
                    "item_name": i.descr,
                    "item_code": i.inventory_item.code,
                    "recv_qty": float(returnable),
                    "rate": float(i.rate),
                    "total": float(returnable * i.rate),
                    "uom": uom_title(i),
                })
            items_json[str(o.pk)] = rows
        context["po_json"] = pos_json
        context["po_items_json"] = items_json
        context["today"] = timezone.localdate()
        return context


class PurchaseReturnQuickCreateView(InventoryManageMixin, View):
    page = "inventory.purchase_returns"
    action = "add"
    def post(self, request):
        po_pk = request.POST.get("purchase_order")
        order = get_object_or_404(PurchaseOrder, pk=po_pk)
        purchase_master = order.purchase_masters.first()
        if not purchase_master:
            messages.error(request, "No purchase master found for this PO.")
            return redirect("inventory:purchase_return_list")
        poi_pks = request.POST.getlist("poi_pk")
        return_qtys = request.POST.getlist("return_qty")
        if not poi_pks:
            messages.error(request, "No items selected.")
            return redirect("inventory:purchase_return_list")
        try:
            with transaction.atomic():
                pr = PurchaseReturnMaster(
                    transaction_id=generate_transaction_id("PRT", PurchaseReturnMaster),
                    purchase_master=purchase_master,
                    return_date=timezone.localdate(),
                    created_by=request.user,
                    updated_by=request.user,
                )
                pr.save()
                for poi_pk, qty_raw in zip(poi_pks, return_qtys):
                    try:
                        qty = Decimal(qty_raw)
                    except Exception:
                        continue
                    if qty <= 0:
                        continue
                    poi = get_object_or_404(PurchaseOrderItem, pk=poi_pk)
                    PurchaseReturnDetail.objects.create(
                        purchase_return_master=pr,
                        inventory_item=poi.inventory_item,
                        quantity=qty,
                        rate=poi.rate,
                        created_by=request.user,
                        updated_by=request.user,
                    )
                post_purchase_return(purchase_return=pr, user=request.user)
        except ValidationError as exc:
            messages.error(request, exc)
            return redirect("inventory:purchase_return_list")
        return redirect(f"{reverse_lazy('inventory:purchase_return_receipt', kwargs={'pk': pr.pk})}?po={po_pk}")


class PurchaseReturnReceiptView(PrintContextMixin, InventoryListMixin, DetailView):
    page = "inventory.purchase_returns"
    model = PurchaseReturnMaster
    template_name = "inventory/purchase_return_receipt.html"
    context_object_name = "pr"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        items = list(self.object.items.select_related("inventory_item").all())
        context["items"] = items
        context["total_qty"] = sum(i.quantity for i in items)
        context["total_return"] = self.object.returned_amount
        context["amount_in_words"] = amount_in_words(self.object.returned_amount)
        context["print_back_url"] = f"{reverse_lazy('inventory:purchase_return_list')}?po={self.request.GET.get('po', '')}"
        return context


class PurchaseReturnCreateView(InventoryManageMixin, CreateView):
    page = "inventory.purchase_returns"
    model = PurchaseReturnMaster
    form_class = PurchaseReturnMasterForm
    template_name = "inventory/simple_form.html"
    success_url = reverse_lazy("inventory:purchase_return_list")
    success_message = "Purchase return saved."
    extra_context = {"title": "Purchase Return"}

    def form_valid(self, form):
        if not form.instance.transaction_id:
            form.instance.transaction_id = generate_transaction_id("PRT", PurchaseReturnMaster)
        return super().form_valid(form)


class PurchaseReturnDetailView(InventoryListMixin, DetailView):
    page = "inventory.purchase_returns"
    model = PurchaseReturnMaster
    template_name = "inventory/purchase_return_detail.html"
    context_object_name = "purchase_return"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["item_form"] = PurchaseReturnDetailForm()
        return context


class PurchaseReturnItemCreateView(InventoryManageMixin, View):
    page = "inventory.purchase_returns"
    action = "add"
    def post(self, request, pk):
        record = get_object_or_404(PurchaseReturnMaster, pk=pk)
        if record.posted == YES:
            messages.error(request, "Posted purchase return cannot be updated.")
            return redirect("inventory:purchase_return_detail", pk=pk)
        form = PurchaseReturnDetailForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.purchase_return_master = record
            item.created_by = request.user
            item.updated_by = request.user
            item.save()
            messages.success(request, "Purchase return item saved.")
        else:
            for errors in form.errors.values():
                for error in errors:
                    messages.error(request, error)
        return redirect("inventory:purchase_return_detail", pk=pk)


class PurchaseReturnPostView(InventoryManageMixin, View):
    page = "inventory.purchase_returns"
    action = "edit"
    def post(self, request, pk):
        record = get_object_or_404(PurchaseReturnMaster, pk=pk)
        try:
            post_purchase_return(purchase_return=record, user=request.user)
            messages.success(request, "Purchase return posted and stock updated.")
        except ValidationError as exc:
            messages.error(request, exc)
        return redirect("inventory:purchase_return_detail", pk=pk)


# ══════════════════════════════════════════════════════════════════════════
# Supplier bills
#
# The screen is built around the goods receipts, not around the order. What is
# billable is what actually arrived and nobody has invoiced yet, so that is
# what the form lists -- and a quantity cannot be typed above it.
# ══════════════════════════════════════════════════════════════════════════


class PurchaseBillListView(InventoryListMixin, ListView):
    """Supplier invoices entered against goods received.

    The tile that matters is GRN Clearing: goods in the godown that no bill has
    been entered for. While it is not zero the payables are understated, and by
    exactly that much.
    """

    page = "inventory.purchase_orders"
    template_name = "inventory/purchase_bill_list.html"
    context_object_name = "bills"
    paginate_by = 25
    queryset = (
        PurchaseBill.objects
        .select_related("supplier", "purchase_order", "created_by")
        .prefetch_related("items__inventory_item")
        .order_by("-bill_date", "-id")
    )
    search_fields = ("bill_num", "supplier_invoice_num", "supplier__name", "purchase_order__purchase_num")
    filter_fields = {"supplier": "supplier_id"}
    date_filters = [{"field": "bill_date", "label": "Bill date"}]

    def get_filter_specs(self):
        supplier_choices = list(Supplier.objects.filter(status=STATUS_ACTIVE).order_by("name").values_list("id", "name"))
        return [{"name": "supplier", "label": "All suppliers", "short_label": "Supplier",
                 "choices": supplier_choices, "value": self.request.GET.get("supplier", "")}]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        pending = billable_receipts()
        context["unbilled_count"] = len(pending)
        context["unbilled_value"] = sum(
            ((row.pending_bill_qty * row.landed_rate).quantize(Decimal("0.01")) for row in pending),
            Decimal("0.00"),
        )
        context["reversal_form"] = ReversalReasonForm()
        carried = self.request.GET.copy()
        for key in ("page",):
            carried.pop(key, None)
        context["base_query"] = carried.urlencode()
        return context


class PurchaseBillCreateView(InventoryManageMixin, View):
    """Enter a supplier's invoice against goods already received.

    Every candidate line carries three figures side by side: what arrived, what
    it was taken into stock at, and what the supplier is asking. Somebody who
    can see all three at once is in a position to notice a rate that was never
    agreed, which is the entire reason the three-way match exists.
    """

    page = "inventory.purchase_orders"
    action = "add"
    template_name = "inventory/purchase_bill_form.html"

    def _candidates(self, supplier=None, order=None):
        rows = billable_receipts(supplier=supplier, purchase_order=order)
        for row in rows:
            row.order_ref = row.purchase_order_item.purchase_order
            row.order_rate = row.purchase_order_item.rate
        return rows

    def get(self, request):
        supplier = None
        supplier_id = (request.GET.get("supplier") or "").strip()
        if supplier_id.isdigit():
            supplier = Supplier.objects.filter(pk=int(supplier_id)).first()

        order = None
        order_id = (request.GET.get("order") or "").strip()
        if order_id.isdigit():
            order = PurchaseOrder.objects.filter(pk=int(order_id)).first()
            supplier = supplier or (order.supplier if order else None)

        form = PurchaseBillForm(initial={
            "supplier": supplier, "bill_date": timezone.localdate(), "supplier_invoice_date": timezone.localdate(),
        })
        return render(request, self.template_name, {
            "title": "Purchase Bill",
            "form": form,
            "candidates": self._candidates(supplier=supplier, order=order),
            "selected_supplier": supplier,
            "selected_order": order,
        })

    def post(self, request):
        form = PurchaseBillForm(request.POST)
        supplier = None
        supplier_id = (request.POST.get("supplier") or "").strip()
        if supplier_id.isdigit():
            supplier = Supplier.objects.filter(pk=int(supplier_id)).first()

        if not form.is_valid():
            for errors in form.errors.values():
                for error in errors:
                    messages.error(request, error)
            return render(request, self.template_name, {
                "title": "Purchase Bill", "form": form,
                "candidates": self._candidates(supplier=supplier), "selected_supplier": supplier,
            })

        # Only the rows that were ticked, and only with the quantity and rate
        # actually typed against them. An untouched row is not a line.
        lines = []
        for receipt_id in request.POST.getlist("receipt_id"):
            if not request.POST.get(f"pick_{receipt_id}"):
                continue
            receipt = PurchaseOrderItemReceived.objects.filter(pk=receipt_id).select_related(
                "purchase_order_item__purchase_order", "inventory_item"
            ).first()
            if not receipt:
                continue
            try:
                quantity = Decimal(request.POST.get(f"qty_{receipt_id}") or "0")
                rate = Decimal(request.POST.get(f"rate_{receipt_id}") or "0")
            except InvalidOperation:
                messages.error(request, f"{receipt.grn_number}: quantity and rate must be numbers.")
                return redirect("inventory:purchase_bill_create")
            lines.append({"receipt": receipt, "quantity": quantity, "rate": rate})

        data = form.cleaned_data
        try:
            bill = create_purchase_bill(
                supplier=data["supplier"],
                supplier_invoice_num=data["supplier_invoice_num"],
                supplier_invoice_date=data.get("supplier_invoice_date"),
                bill_date=data["bill_date"],
                due_date=data.get("due_date"),
                lines=lines,
                freight_amount=data.get("freight_amount") or Decimal("0"),
                discount_amount=data.get("discount_amount") or Decimal("0"),
                tax_amount=data.get("tax_amount") or Decimal("0"),
                remarks=data.get("remarks") or "",
                variance_approved=bool(data.get("variance_approved")),
                user=request.user,
            )
        except ValidationError as exc:
            for message in exc.messages:
                messages.error(request, message)
            return render(request, self.template_name, {
                "title": "Purchase Bill", "form": form,
                "candidates": self._candidates(supplier=data["supplier"]), "selected_supplier": data["supplier"],
            })

        messages.success(
            request,
            f"Bill {bill.bill_num} posted for {bill.total_amount}. "
            "GRN clearing released, payable created, input tax claimable.",
        )
        return redirect("inventory:purchase_bill_detail", pk=bill.pk)


class PurchaseBillDetailView(InventoryListMixin, DetailView):
    page = "inventory.purchase_orders"
    model = PurchaseBill
    template_name = "inventory/purchase_bill_detail.html"
    context_object_name = "bill"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        ok, why = can_reverse_bill(self.object)
        context["can_reverse"] = ok
        context["cannot_reverse_because"] = why
        context["reversal_form"] = ReversalReasonForm()
        return context


class PurchaseBillReverseView(InventoryManageMixin, View):
    """Withdraw a posted bill: value back into GRN Clearing, payable off."""

    page = "inventory.purchase_orders"
    action = "reverse"

    def post(self, request, pk):
        bill = get_object_or_404(PurchaseBill, pk=pk)
        form = ReversalReasonForm(request.POST)
        if not form.is_valid():
            messages.error(request, "A reversal needs a reason.")
            return redirect("inventory:purchase_bill_detail", pk=pk)
        try:
            mirror = reverse_purchase_bill(bill=bill, reason=form.cleaned_data["reason"], user=request.user)
            messages.success(
                request,
                f"{bill.bill_num} reversed by {mirror.bill_num}. The receipts are unbilled again, "
                "so the correct invoice can be entered against them.",
            )
        except ValidationError as exc:
            messages.error(request, "; ".join(exc.messages))
        return redirect("inventory:purchase_bill_detail", pk=pk)


class GoodsReceiptCreateView(InventoryManageMixin, View):
    """Book a delivery in against an order, as one document.

    The register already lets a row be expanded and received in place, and an
    order's own page has a panel for it. This is the third way in, and it is the
    one for the store: somebody standing at the gate with a delivery in front of
    them, who knows the supplier and the vehicle but has not got a particular
    row on a list in mind.

    It receives several lines of one order in a single pass, which is what a
    delivery actually is -- so the freight paid on it can be split across those
    lines by value rather than being typed against one of them.
    """

    page = "inventory.grn"
    action = "edit"
    template_name = "inventory/goods_receipt_form.html"

    def _open_orders(self, supplier=None):
        """Orders goods can still be booked in against.

        Draft orders are left out: nobody has committed to them, so nothing
        should be arriving. So are the ones cancelled or closed short, where the
        business has already said the goods are not coming.
        """
        rows = (
            PurchaseOrder.objects
            .filter(is_direct=False, status__in=(STATUS_RAISED, STATUS_PARTIAL_RECEIVED))
            .select_related("supplier")
            .prefetch_related("items__inventory_item", "items__uom")
            .order_by("-purchase_date", "-id")
        )
        if supplier is not None:
            rows = rows.filter(supplier=supplier)
        picked = []
        for order in rows:
            # The lines still owed on this order, hung off it so the picker can
            # show what is actually outstanding without a second query per row.
            order.open_lines = [line for line in order.items.all() if line.open_receive_qty > 0]
            if not order.open_lines:
                continue
            # Totals over the whole order, not only the lines still open, so the
            # picker states how far along the order is rather than what is left.
            lines = list(order.items.all())
            order.ordered_qty = sum((line.quantity + line.extra_qty for line in lines), Decimal("0.0000"))
            order.received_qty = sum((line.total_receive_qty for line in lines), Decimal("0.0000"))
            order.balance_qty = sum((line.open_receive_qty for line in lines), Decimal("0.0000"))
            order.order_amount = sum((line.total_amount for line in lines), Decimal("0.00"))
            order.item_count = len(lines)
            picked.append(order)
        return picked

    def _context(self, request, supplier=None, order=None):
        # Kept across the supplier/order reload, which is a GET, so a date typed
        # before the order was picked is not quietly thrown away.
        grn_date = (request.GET.get("grn_date") or "").strip()
        orders = self._open_orders(supplier=supplier)
        if order is not None and order not in orders:
            # Somebody arrived on a link to an order that has since been
            # completed or closed. Say so rather than showing an empty grid.
            order = None
        lines = []
        if order is not None:
            for line in order.items.all():
                if line.open_receive_qty <= 0:
                    continue
                line.max_now = line.open_receive_qty
                lines.append(line)
        return {
            "title": "Goods Receipt Note",
            "suppliers": Supplier.objects.filter(status=STATUS_ACTIVE).order_by("name"),
            "orders": orders,
            "selected_supplier": supplier,
            "selected_order": order,
            "lines": lines,
            "today": timezone.localdate(),
            "next_grn_no": next_grn_number(),
            "grn_date": grn_date,
            "clearing_balance": -(balance_of_grn_clearing()),
        }

    def get(self, request):
        supplier = Supplier.objects.filter(pk=request.GET.get("supplier") or 0).first()
        order = PurchaseOrder.objects.filter(pk=request.GET.get("order") or 0).first()
        if order and not supplier:
            supplier = order.supplier
        return render(request, self.template_name, self._context(request, supplier=supplier, order=order))

    def post(self, request):
        supplier = Supplier.objects.filter(pk=request.POST.get("supplier") or 0).first()
        order = PurchaseOrder.objects.filter(pk=request.POST.get("order") or 0).first()
        if not order:
            messages.error(request, "Choose the purchase order these goods arrived against.")
            return redirect("inventory:goods_receipt_create")

        receive_date = request.POST.get("receive_date") or str(timezone.localdate())
        # One delivery, one GRN number: every line booked in this pass carries it.
        grn_number = (request.POST.get("grn_number") or "").strip()
        rv_number = " ".join(part for part in (
            (request.POST.get("dc_number") or "").strip(),
            (request.POST.get("vehicle") or "").strip(),
        ) if part)[:80]
        # What the store saw, kept with the receipt rather than in somebody's
        # head: who checked it, and anything they want on the record.
        inspected = (request.POST.get("inspected_by") or "").strip()
        note = (request.POST.get("remarks") or "").strip()
        remarks = " — ".join(part for part in (note, f"Inspected by {inspected}" if inspected else "") if part)

        try:
            freight_total = Decimal(request.POST.get("freight") or "0")
        except InvalidOperation:
            freight_total = Decimal("0")

        # Gather first, post second: freight is split across the lines by their
        # value, which cannot be worked out until every line is known.
        picked = []
        for line in order.items.all():
            try:
                quantity = Decimal(request.POST.get(f"qty_{line.pk}") or "0")
                rejected = Decimal(request.POST.get(f"rej_{line.pk}") or "0")
                rate = Decimal(request.POST.get(f"rate_{line.pk}") or "0")
            except InvalidOperation:
                messages.error(request, f"{line.descr}: quantity, rejected and rate must be numbers.")
                return redirect(f"{reverse_lazy('inventory:goods_receipt_create')}?order={order.pk}")
            accepted = quantity - rejected
            if accepted <= 0:
                continue
            picked.append((line, accepted, rejected, rate or line.rate or Decimal("0")))

        if not picked:
            messages.error(request, "Nothing was accepted — enter a received quantity on at least one line.")
            return redirect(f"{reverse_lazy('inventory:goods_receipt_create')}?order={order.pk}")

        shares = apportion_freight(freight_total, [accepted * rate for _l, accepted, _r, rate in picked])

        receipt_pks = []
        try:
            with transaction.atomic():
                for (line, accepted, rejected, rate), freight in zip(picked, shares):
                    line_note = remarks
                    if rejected > 0:
                        # A rejection is a fact about the delivery and belongs on
                        # the receipt. It is not taken into stock, so it appears
                        # nowhere else at all unless it is written down here.
                        line_note = f"{line_note} — {rejected} rejected" if line_note else f"{rejected} rejected"
                    receipt = receive_purchase_order_item(
                        purchase_order_item=line,
                        quantity=accepted,
                        extra_qty=Decimal("0"),
                        retail_price=rate,
                        receive_date=receive_date,
                        invoice_num="",
                        invoice_date=None,
                        rv_number=rv_number,
                        remarks=line_note,
                        user=request.user,
                        freight_amount=freight,
                        grn_number=grn_number,
                    )
                    receipt_pks.append(str(receipt.pk))
        except ValidationError as exc:
            for message in exc.messages:
                messages.error(request, message)
            return redirect(f"{reverse_lazy('inventory:goods_receipt_create')}?order={order.pk}")

        messages.success(
            request,
            f"Goods receipt posted for {order.purchase_num}: {len(receipt_pks)} line(s) into stock. "
            "The value waits in GRN clearing until the supplier's bill is entered.",
        )
        return redirect(f"{reverse_lazy('inventory:grn_print', kwargs={'pk': order.pk})}?receipts={','.join(receipt_pks)}")


def balance_of_grn_clearing():
    """What the GRN clearing account is holding right now.

    Read straight off the posted voucher lines rather than from a stored total,
    so the figure on the screen cannot drift from the ledger it claims to show.
    """
    from apps.core.constants import GL_GRN_CLEARING_PATH
    from apps.finance.services import gl_account

    account = gl_account(GL_GRN_CLEARING_PATH)
    rows = AccountVoucherLine.objects.filter(account_no=account.code).aggregate(
        debit=Sum("debit_amount"), credit=Sum("credit_amount")
    )
    return (rows["debit"] or Decimal("0")) - (rows["credit"] or Decimal("0"))
